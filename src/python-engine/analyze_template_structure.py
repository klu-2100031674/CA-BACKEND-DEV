"""
Analyze Excel template structure to identify input vs formula cells.
This script inspects any Excel template and generates the correct cell mapping.
"""

import openpyxl
import sys
import json
from pathlib import Path

def analyze_excel_template(excel_path):
    """
    Analyze Excel template to identify:
    - Which cells contain formulas (should NOT be overwritten)
    - Which cells are input cells (safe to write)
    - Column structure for each section
    """
    
    print(f"\n{'='*80}")
    print(f"ANALYZING TEMPLATE: {Path(excel_path).name}")
    print(f"{'='*80}\n")
    
    workbook = openpyxl.load_workbook(excel_path, data_only=False)
    
    # Focus on Assumptions.1 sheet (main data sheet)
    if 'Assumptions.1' not in workbook.sheetnames:
        print(f"ERROR: 'Assumptions.1' sheet not found!")
        print(f"Available sheets: {workbook.sheetnames}")
        return None
    
    sheet = workbook['Assumptions.1']
    
    # Analyze structure
    analysis = {
        'template_name': Path(excel_path).name,
        'sections': {},
        'input_columns': set(),
        'formula_columns': set(),
        'cell_mappings': {}
    }
    
    # Scan rows 1-200 to find all sections
    print("🔍 SCANNING EXCEL STRUCTURE (Rows 1-200)...\n")
    
    current_section = None
    section_start_row = None
    
    for row in range(1, 201):
        # Check columns G, H, I, J for content
        col_g = sheet[f'G{row}'].value
        col_h = sheet[f'H{row}'].value
        col_i = sheet[f'I{row}'].value
        
        # Detect section headers (usually in column G)
        if col_g and isinstance(col_g, str) and len(str(col_g)) > 3:
            # Potential section header
            if current_section:
                # Save previous section
                analysis['sections'][current_section] = {
                    'start_row': section_start_row,
                    'end_row': row - 1
                }
            
            current_section = str(col_g).strip()
            section_start_row = row
            print(f"📌 Found Section: {current_section} (Row {row})")
        
        # Check for formulas in columns H, I, J
        for col_letter in ['G', 'H', 'I', 'J']:
            cell = sheet[f'{col_letter}{row}']
            if cell.value is not None:
                if hasattr(cell, 'value') and isinstance(cell.value, str) and cell.value.startswith('='):
                    # Formula cell
                    analysis['formula_columns'].add(col_letter)
                    analysis['cell_mappings'][f'{col_letter.lower()}{row}'] = {
                        'type': 'formula',
                        'formula': cell.value,
                        'section': current_section
                    }
                elif col_letter in ['H', 'I', 'J'] and cell.value is not None:
                    # Potential input cell (has value but not formula)
                    analysis['input_columns'].add(col_letter)
                    analysis['cell_mappings'][f'{col_letter.lower()}{row}'] = {
                        'type': 'input',
                        'value': cell.value,
                        'section': current_section
                    }
    
    # Save last section
    if current_section:
        analysis['sections'][current_section] = {
            'start_row': section_start_row,
            'end_row': 200
        }
    
    print(f"\n{'='*80}")
    print("📊 ANALYSIS RESULTS")
    print(f"{'='*80}\n")
    
    print(f"Total Sections Found: {len(analysis['sections'])}")
    for section_name, section_info in analysis['sections'].items():
        print(f"  • {section_name}: Rows {section_info['start_row']}-{section_info['end_row']}")
    
    print(f"\nColumns Used:")
    print(f"  • Input Columns: {sorted(analysis['input_columns'])}")
    print(f"  • Formula Columns: {sorted(analysis['formula_columns'])}")
    
    # Generate section-specific mappings
    print(f"\n{'='*80}")
    print("📋 SECTION-SPECIFIC CELL MAPPINGS")
    print(f"{'='*80}\n")
    
    section_mappings = {}
    for section_name, section_info in analysis['sections'].items():
        input_rows = []
        formula_rows = []
        
        for row in range(section_info['start_row'], section_info['end_row'] + 1):
            # Check primary data column (usually I)
            cell_ref = f'i{row}'
            if cell_ref in analysis['cell_mappings']:
                if analysis['cell_mappings'][cell_ref]['type'] == 'input':
                    input_rows.append(row)
                elif analysis['cell_mappings'][cell_ref]['type'] == 'formula':
                    formula_rows.append(row)
        
        if input_rows or formula_rows:
            section_mappings[section_name] = {
                'input_rows': input_rows,
                'formula_rows': formula_rows,
                'row_range': f"{section_info['start_row']}-{section_info['end_row']}"
            }
            
            print(f"📦 {section_name}")
            print(f"   Range: Rows {section_info['start_row']}-{section_info['end_row']}")
            if input_rows:
                print(f"   ✅ Input Rows ({len(input_rows)}): {input_rows[:10]}{'...' if len(input_rows) > 10 else ''}")
            if formula_rows:
                print(f"   ⚠️  Formula Rows ({len(formula_rows)}): {formula_rows[:10]}{'...' if len(formula_rows) > 10 else ''}")
            print()
    
    # Check for H column inputs (like CC2)
    print(f"{'='*80}")
    print("🔎 CHECKING FOR H COLUMN INPUTS (Alternative Input Column)")
    print(f"{'='*80}\n")
    
    h_column_inputs = []
    for row in range(1, 201):
        cell = sheet[f'H{row}']
        if cell.value is not None and not (isinstance(cell.value, str) and cell.value.startswith('=')):
            # H column has value and it's not a formula
            i_cell = sheet[f'I{row}']
            if i_cell.value and isinstance(i_cell.value, str) and i_cell.value.startswith('='):
                # I column has formula, H column is input
                h_column_inputs.append({
                    'row': row,
                    'h_value': cell.value,
                    'i_formula': i_cell.value
                })
    
    if h_column_inputs:
        print(f"Found {len(h_column_inputs)} H-column input cells:")
        for item in h_column_inputs[:15]:
            print(f"  Row {item['row']}: H{item['row']}={item['h_value']} → I{item['row']}={item['i_formula']}")
        if len(h_column_inputs) > 15:
            print(f"  ... and {len(h_column_inputs) - 15} more")
    else:
        print("No H-column inputs found (all inputs are in I column)")
    
    # Generate JavaScript mapping code
    print(f"\n{'='*80}")
    print("💻 GENERATED JAVASCRIPT MAPPING (for React form)")
    print(f"{'='*80}\n")
    
    for section_name, mapping in section_mappings.items():
        safe_section_name = section_name.replace(' ', '').replace('-', '').replace('/', '')
        print(f"// {section_name}")
        print(f"const {safe_section_name}InputRows = {mapping['input_rows']};")
        print()
    
    # Save to JSON file
    output_file = Path(excel_path).stem + '_mapping.json'
    output_path = Path(__file__).parent / output_file
    
    with open(output_path, 'w') as f:
        json.dump({
            'template_name': analysis['template_name'],
            'sections': section_mappings,
            'h_column_inputs': [item['row'] for item in h_column_inputs]
        }, f, indent=2)
    
    print(f"\n✅ Mapping saved to: {output_file}")
    
    return analysis

if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Usage: python analyze_template_structure.py <path_to_excel_template>")
        print("\nExample:")
        print("  python analyze_template_structure.py ../../templates/excel/format CC2.xlsx")
        sys.exit(1)
    
    excel_path = sys.argv[1]
    
    if not Path(excel_path).exists():
        print(f"ERROR: File not found: {excel_path}")
        sys.exit(1)
    
    analyze_excel_template(excel_path)
