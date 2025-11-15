"""
Inspect format CC2.xlsx to see actual cell structure
"""
import openpyxl
from pathlib import Path

# Load the workbook
wb_path = Path(__file__).parent.parent.parent / 'templates' / 'excel' / 'format CC2.xlsx'
wb = openpyxl.load_workbook(wb_path, data_only=False)

# Check Assumptions.1 sheet
if 'Assumptions.1' in wb.sheetnames:
    ws = wb['Assumptions.1']
    print("=" * 80)
    print("ASSUMPTIONS.1 SHEET - Key Cell Mappings")
    print("=" * 80)
    
    # Check rows 3-20 for General Info, Finance, Years
    print("\n--- GENERAL INFO & FINANCE (Rows 3-20) ---")
    for row in range(3, 21):
        cell_i = ws[f'I{row}']
        cell_h = ws[f'H{row}']
        cell_g = ws[f'G{row}']
        
        # Print row number and values/formulas in columns G, H, I
        print(f"Row {row:2d}:")
        if cell_g.value:
            print(f"  G{row}: {str(cell_g.value)[:60]}")
        if cell_h.value:
            val_h = cell_h.value if not str(cell_h.value).startswith('=') else f"Formula: {str(cell_h.value)[:50]}"
            print(f"  H{row}: {val_h}")
        if cell_i.value:
            val_i = cell_i.value if not str(cell_i.value).startswith('=') else f"Formula: {str(cell_i.value)[:50]}"
            print(f"  I{row}: {val_i}")
    
    print("\n--- AUDITED STATEMENTS (Sample rows 22-30) ---")
    for row in range(22, 31):
        cell_i = ws[f'I{row}']
        cell_h = ws[f'H{row}']
        cell_g = ws[f'G{row}']
        
        print(f"Row {row:2d}:")
        if cell_g.value:
            print(f"  G{row}: {str(cell_g.value)[:60]}")
        if cell_h.value:
            val_h = cell_h.value if not str(cell_h.value).startswith('=') else f"Formula: {str(cell_h.value)[:50]}"
            print(f"  H{row}: {val_h}")
        if cell_i.value:
            val_i = cell_i.value if not str(cell_i.value).startswith('=') else f"Formula: {str(cell_i.value)[:50]}"
            print(f"  I{row}: {val_i}")

    print("\n--- ASSUMPTIONS (Sample rows 67-75) ---")
    for row in range(67, 76):
        cell_i = ws[f'I{row}']
        cell_h = ws[f'H{row}']
        cell_g = ws[f'G{row}']
        
        print(f"Row {row:2d}:")
        if cell_g.value:
            print(f"  G{row}: {str(cell_g.value)[:60]}")
        if cell_h.value:
            val_h = cell_h.value if not str(cell_h.value).startswith('=') else f"Formula: {str(cell_h.value)[:50]}"
            print(f"  H{row}: {val_h}")
        if cell_i.value:
            val_i = cell_i.value if not str(cell_i.value).startswith('=') else f"Formula: {str(cell_i.value)[:50]}"
            print(f"  I{row}: {val_i}")

else:
    print("Assumptions.1 sheet not found!")
    print("Available sheets:", wb.sheetnames)

wb.close()
