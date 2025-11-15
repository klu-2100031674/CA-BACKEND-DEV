"""
Analyze all CC Format Excel files to identify input fields and assumptions
This script helps determine what frontend forms need to be created for CC2-CC6
"""

import openpyxl
from openpyxl import load_workbook
import json
import sys
from pathlib import Path
import re

def analyze_excel_file(file_path, format_name):
    """Analyze an Excel file to find input cells and assumptions"""
    print(f"\n{'='*80}")
    print(f"ANALYZING: {format_name}")
    print(f"File: {file_path}")
    print(f"{'='*80}")
    
    try:
        wb = load_workbook(file_path, data_only=False)
        
        # Get all sheet names
        print(f"\n📋 SHEETS ({len(wb.sheetnames)}):")
        for i, sheet_name in enumerate(wb.sheetnames, 1):
            print(f"   {i}. {sheet_name}")
        
        # Focus on Assumptions sheet (usually the input sheet)
        assumption_sheets = [s for s in wb.sheetnames if 'assumption' in s.lower() or 'input' in s.lower()]
        
        if not assumption_sheets:
            print(f"\n⚠️  No Assumptions sheet found. Analyzing first sheet: {wb.sheetnames[0]}")
            assumption_sheets = [wb.sheetnames[0]]
        
        for sheet_name in assumption_sheets:
            print(f"\n🔍 ANALYZING SHEET: {sheet_name}")
            print(f"{'-'*80}")
            
            ws = wb[sheet_name]
            
            # Find input cells (cells that appear to be user inputs)
            # These typically have:
            # 1. A label in column to the left or above
            # 2. No formula (just values)
            # 3. Often in columns I, J, or later
            
            input_fields = []
            
            # Scan reasonable range (rows 1-200, cols A-M)
            for row in range(1, 201):
                for col in range(1, 14):  # A to M
                    cell = ws.cell(row=row, column=col)
                    
                    # Skip empty cells
                    if cell.value is None:
                        continue
                    
                    # Look for potential label cells (text in columns A-H)
                    if col <= 8 and isinstance(cell.value, str):
                        # Check if next cell (same row) might be input
                        next_cell = ws.cell(row=row, column=col + 1)
                        
                        # Check multiple columns to the right
                        for offset in range(1, 6):
                            input_cell = ws.cell(row=row, column=col + offset)
                            
                            if input_cell.value is not None:
                                # Determine if it's likely an input field
                                has_formula = input_cell.data_type == 'f'
                                is_in_input_column = (col + offset) >= 9  # Column I or later
                                
                                # If not a formula or in input columns, it might be an input
                                if not has_formula or is_in_input_column:
                                    cell_ref = f"{get_column_letter(col + offset)}{row}"
                                    
                                    input_fields.append({
                                        'label': str(cell.value).strip()[:100],
                                        'cell': cell_ref,
                                        'value': input_cell.value,
                                        'has_formula': has_formula,
                                        'type': type(input_cell.value).__name__
                                    })
                                    break  # Found input for this label
            
            # Remove duplicates and sort
            unique_fields = {}
            for field in input_fields:
                key = field['cell']
                if key not in unique_fields:
                    unique_fields[key] = field
            
            input_fields = list(unique_fields.values())
            input_fields.sort(key=lambda x: (int(re.search(r'\d+', x['cell']).group()), x['cell']))
            
            print(f"\n📝 FOUND {len(input_fields)} POTENTIAL INPUT FIELDS:")
            print(f"{'-'*80}")
            
            # Group by section (every 10 rows)
            current_section = 0
            for field in input_fields[:100]:  # Limit to first 100 for readability
                row_num = int(re.search(r'\d+', field['cell']).group())
                section = row_num // 10
                
                if section != current_section:
                    current_section = section
                    print(f"\n  --- Rows {section * 10} - {(section + 1) * 10 - 1} ---")
                
                # Format value for display
                val_str = str(field['value'])[:40]
                if len(str(field['value'])) > 40:
                    val_str += '...'
                
                formula_indicator = ' [F]' if field['has_formula'] else ''
                print(f"  {field['cell']:6} | {field['label'][:50]:50} | {val_str:40} {formula_indicator}")
        
        # Summary
        print(f"\n{'='*80}")
        print(f"SUMMARY FOR {format_name}:")
        print(f"  - Total Sheets: {len(wb.sheetnames)}")
        print(f"  - Assumption Sheets: {len(assumption_sheets)}")
        print(f"  - Potential Input Fields: {len(input_fields)}")
        print(f"{'='*80}\n")
        
        return {
            'format_name': format_name,
            'total_sheets': len(wb.sheetnames),
            'sheet_names': wb.sheetnames,
            'assumption_sheets': assumption_sheets,
            'input_fields': input_fields[:50]  # First 50 fields
        }
        
    except Exception as e:
        print(f"❌ ERROR analyzing {format_name}: {str(e)}")
        import traceback
        traceback.print_exc()
        return None


def get_column_letter(col_num):
    """Convert column number to letter (1->A, 2->B, etc.)"""
    string = ""
    while col_num > 0:
        col_num, remainder = divmod(col_num - 1, 26)
        string = chr(65 + remainder) + string
    return string


def main():
    # Path to Excel templates
    templates_path = Path(__file__).parent.parent.parent / 'templates' / 'excel'
    
    # Excel files to analyze
    excel_files = [
        ('format CC1.xlsx', 'Format CC1'),
        ('format CC2.xlsx', 'Format CC2'),
        ('Format CC3.xlsx', 'Format CC3'),
        ('format CC4.xlsx', 'Format CC4'),
        ('format CC5.xlsx', 'Format CC5'),
        ('format CC6.xlsx', 'Format CC6'),
    ]
    
    all_results = []
    
    for filename, format_name in excel_files:
        file_path = templates_path / filename
        if file_path.exists():
            result = analyze_excel_file(str(file_path), format_name)
            if result:
                all_results.append(result)
        else:
            print(f"⚠️  File not found: {file_path}")
    
    # Save results to JSON
    output_path = templates_path.parent / 'analysis_results.json'
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(all_results, f, indent=2, default=str)
    
    print(f"\n✅ Analysis complete! Results saved to: {output_path}")
    print(f"\nAnalyzed {len(all_results)} formats successfully.")


if __name__ == '__main__':
    main()
