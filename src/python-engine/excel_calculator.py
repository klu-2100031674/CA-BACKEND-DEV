"""Excel payload applier.

Updates cells based on the incoming payload, leaves the source template untouched,
and writes an updated copy to the temp directory for verification.
"""

import json
import datetime
import os
import sys
import re
import uuid
from typing import Any, Dict, List, Optional
import base64

import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_to_tuple
import pandas as pd
import numpy as np
from pathlib import Path
import subprocess
import shutil


def _normalized_sheet_key(value: str) -> str:
    """Normalize sheet names/inputs for loose matching.

    Used in include/exclude filtering and stable filenames.
    """
    return re.sub(r"[\s_\-]+", "", (value or "").strip().lower())


def _get_safe_temp_dir() -> str:
    """Return a temp directory that works on the current OS.

    Render/Linux deployments sometimes receive a Windows-style TEMP_DIR (e.g. D:\\...).
    That breaks path operations inside containers, so we fall back to the system temp.
    """
    import tempfile

    candidate = os.getenv("TEMP_DIR")
    if not candidate:
        return tempfile.gettempdir()

    # If we're not on Windows and the path looks like a Windows drive path, ignore it.
    if os.name != "nt" and re.match(r"^[A-Za-z]:\\", candidate):
        return tempfile.gettempdir()

    return candidate

# Configuration Flag
# Prefer env var so Render/Docker can control behavior.
_use_com_env = os.getenv("USE_COM_INTERFACE", "False").strip().lower()
USE_COM_INTERFACE = _use_com_env in {"1", "true", "yes", "y", "on"}

# Windows COM for Excel automation
try:
    import win32com.client
    import pythoncom
    COM_AVAILABLE = True
except ImportError:
    COM_AVAILABLE = False
    print("Warning: pywin32 not available. PDF generation will fallback to LibreOffice/Linux mode.", file=sys.stderr)

# If COM libs aren't available, force No-COM mode regardless of env.
USE_COM_INTERFACE = bool(USE_COM_INTERFACE and COM_AVAILABLE)

# Best-practice for Linux PDF export: flatten formulas to values before LibreOffice touches the file.
_freeze_env = os.getenv("FREEZE_FORMULAS_TO_VALUES", "True").strip().lower()
FREEZE_FORMULAS_TO_VALUES = _freeze_env in {"1", "true", "yes", "y", "on"}

import subprocess
from openpyxl import load_workbook
import shutil

try:
    from xlcalculator import ModelCompiler
    from xlcalculator.evaluator import Evaluator
    from xlcalculator.xlfunctions import func_xltypes, xlerrors, xl
    XL_CALC_AVAILABLE = True
except ImportError:
    XL_CALC_AVAILABLE = False

_XL_CUSTOM_FUNCS_REGISTERED = False


def _freeze_workbook_formulas_to_values(source_xlsx_path: str, output_xlsx_path: str) -> Dict[str, int]:
    """Create a copy of the workbook with formulas replaced by cached values.

    This prevents LibreOffice from re-evaluating formulas (which can yield #NAME? for
    unsupported Excel functions) during PDF conversion.

    Note: This uses the cached results stored in the XLSX file. If the workbook was
    not recalculated after updates, some cached results may be missing.
    """

    converted = 0
    missing = 0
    skipped_sheets = 0

    wb_formula = load_workbook(source_xlsx_path, data_only=False, keep_vba=True)
    wb_values = load_workbook(source_xlsx_path, data_only=True, keep_vba=True)

    try:
        for sheet_name in wb_formula.sheetnames:
            if sheet_name not in wb_values.sheetnames:
                skipped_sheets += 1
                continue

            ws_formula = wb_formula[sheet_name]
            ws_values = wb_values[sheet_name]

            for row in ws_formula.iter_rows():
                for cell in row:
                    cell_value = cell.value
                    is_formula = cell.data_type == "f" or (
                        isinstance(cell_value, str) and cell_value.startswith("=")
                    )
                    if not is_formula:
                        continue

                    cached_value = ws_values[cell.coordinate].value
                    if cached_value is None:
                        missing += 1
                        continue

                    cell.value = cached_value
                    converted += 1

        # Avoid recalculation prompts by consumers
        try:
            wb_formula.calculation.fullCalcOnLoad = False
        except Exception:
            pass

        wb_formula.save(output_xlsx_path)
    finally:
        try:
            wb_values.close()
        except Exception:
            pass
        try:
            wb_formula.close()
        except Exception:
            pass

    return {
        "converted": converted,
        "missing_cached": missing,
        "skipped_sheets": skipped_sheets,
    }


def _estimate_used_rows_and_cols(ws) -> tuple[int, int]:
    """Best-effort used range estimation for page-fitting decisions."""
    try:
        dim = ws.calculate_dimension()  # e.g. A1:K80
        if not dim or ":" not in dim:
            return (ws.max_row or 1, ws.max_column or 1)
        end = dim.split(":", 1)[1]
        row, col = coordinate_to_tuple(end)
        return (row or 1, col or 1)
    except Exception:
        return (ws.max_row or 1, ws.max_column or 1)


def _apply_linux_page_setup(ws, sheet_name: str) -> None:
    """Apply COM-like page setup so LibreOffice exports match formatting closely."""
    norm_name = normalize_sheet_name(sheet_name)
    designed = norm_name in {"coverpage", "cover page", "cover", "index"}

    used_rows, used_cols = _estimate_used_rows_and_cols(ws)

    # Orientation rules similar to COM
    wide_sheets = {
        "workings for pl & bs",
        "workings for plbs",
        "workings for pl bs 2",
        "depreciation",
        "repayment schedule",
        "working capital loan",
    }
    is_wide = used_cols > 10 or any(key in norm_name for key in wide_sheets)

    # A4 + fit
    try:
        ws.page_setup.paperSize = 9  # A4
    except Exception:
        pass

    try:
        ws.page_setup.orientation = "landscape" if is_wide else "portrait"
    except Exception:
        pass

    # Margins (in inches)
    try:
        ws.page_margins.left = 0.25
        ws.page_margins.right = 0.25
        ws.page_margins.top = 0.40 if designed else 0.30
        ws.page_margins.bottom = 0.40 if designed else 0.30
        ws.page_margins.header = 0.20
        ws.page_margins.footer = 0.20
    except Exception:
        pass

    # Centering
    try:
        ws.page_setup.horizontalCentered = True
        ws.page_setup.verticalCentered = bool(designed)
    except Exception:
        pass

    # Fit-to-page settings
    try:
        ws.sheet_properties.pageSetUpPr.fitToPage = True
    except Exception:
        pass

    try:
        ws.page_setup.fitToWidth = 1
        if designed:
            ws.page_setup.fitToHeight = 1
        else:
            pages_tall = max(1, (used_rows + 59) // 60)
            ws.page_setup.fitToHeight = pages_tall
    except Exception:
        pass

    # Respect existing PrintArea when defined (helps exclude side instructions)
    # We intentionally do NOT clear/override it here.


def _generate_pdf_via_libreoffice(excel_path: str, sheet_name: str, output_path: str) -> bool:
    """Generate a single-sheet PDF using LibreOffice by isolating the sheet."""
    temp_processing_dir = os.path.join(_get_safe_temp_dir(), "temp_libre_single")
    os.makedirs(temp_processing_dir, exist_ok=True)

    run_id = uuid.uuid4().hex[:8]
    temp_excel_name = f"temp_{run_id}_{_normalized_sheet_key(sheet_name)}.xlsx"
    temp_excel_path = os.path.join(temp_processing_dir, temp_excel_name)
    out_dir = os.path.dirname(os.path.abspath(output_path))
    os.makedirs(out_dir, exist_ok=True)

    try:
        shutil.copy2(excel_path, temp_excel_path)

        wb_temp = load_workbook(temp_excel_path)
        actual_sheet = find_sheet_match(sheet_name, wb_temp.sheetnames) or sheet_name
        if actual_sheet not in wb_temp.sheetnames:
            wb_temp.close()
            return False

        # Delete other sheets to force one-sheet export
        for s in list(wb_temp.sheetnames):
            if s != actual_sheet:
                del wb_temp[s]

        _apply_linux_page_setup(wb_temp[actual_sheet], actual_sheet)
        wb_temp.save(temp_excel_path)
        wb_temp.close()

        cmd = [
            "soffice",
            "--headless",
            "--convert-to",
            "pdf",
            "--outdir",
            out_dir,
            temp_excel_path,
        ]
        print(f"[PDF Generator] LibreOffice running: {' '.join(cmd)}", file=sys.stderr)
        result = subprocess.run(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, timeout=90)
        if result.returncode != 0:
            print(f"[PDF Generator] LibreOffice error: {result.stderr.decode(errors='replace')}", file=sys.stderr)
            return False

        generated_pdf_name = temp_excel_name.replace(".xlsx", ".pdf")
        generated_pdf_path = os.path.join(out_dir, generated_pdf_name)
        if not os.path.exists(generated_pdf_path):
            print(f"[PDF Generator] LibreOffice output PDF missing: {generated_pdf_path}", file=sys.stderr)
            return False

        if os.path.exists(output_path):
            try:
                os.remove(output_path)
            except Exception:
                pass
        shutil.move(generated_pdf_path, output_path)
        return os.path.exists(output_path)
    finally:
        try:
            if os.path.exists(temp_excel_path):
                os.remove(temp_excel_path)
        except Exception:
            pass
        try:
            shutil.rmtree(temp_processing_dir)
        except Exception:
            pass


def get_final_sheet_name(template_name: str) -> str:
    """
    Get the correct 'Final workings' sheet name based on template type.
    
    Args:
        template_name: The template file name or identifier
        
    Returns:
        The exact sheet name to use for this template
    """
    template_upper = template_name.upper()
    
    # CC1 -> FinalWorkings
    if 'CC1' in template_upper or 'FORMAT CC1' in template_upper:
        return 'FinalWorkings'
    # CC2 -> FinalWorkings
    elif 'CC2' in template_upper or 'FORMAT CC2' in template_upper:
        return 'FinalWorkings'
    # CC3 -> Finalworkings (lowercase 'w')
    elif 'CC3' in template_upper or 'FORMAT CC3' in template_upper:
        return 'Finalworkings'
    # CC4 -> Finalworkings (lowercase 'w')
    elif 'CC4' in template_upper or 'FORMAT CC4' in template_upper:
        return 'Finalworkings'
    # CC5 -> FinalWorkings
    elif 'CC5' in template_upper or 'FORMAT CC5' in template_upper:
        return 'FinalWorkings'
    # CC6 -> Final workings (with space)
    elif 'CC6' in template_upper or 'FORMAT CC6' in template_upper:
        return 'Final workings'
    # Term Loan -> Final workings (with space)
    elif 'TERM LOAN' in template_upper or 'TERM_LOAN' in template_upper:
        return 'Final workings'
    # Default fallback
    else:
        return 'Finalworkings'


def extract_sheet_data_with_com(excel_path: str, sheet_name: str = None) -> List[Dict]:
    """
    Extract sheet data using Excel COM automation for accurate cell values, formulas, and formatting.
    This method preserves the exact Excel structure including merged cells and formatting.
    
    Args:
        excel_path: Path to the Excel file
        sheet_name: Specific sheet to extract (if None, extracts all sheets)
        
    Returns:
        List of sheet objects in Luckysheet format with complete formatting
    """
    if not COM_AVAILABLE:
        print("[COM Extraction] COM not available, falling back to openpyxl", file=sys.stderr)
        return None
    
    try:
        import win32com.client
        import pythoncom
        
        pythoncom.CoInitialize()
        # Use DispatchEx to ensure a new independent Excel instance for each process
        excel_app = win32com.client.DispatchEx("Excel.Application")
        try:
            excel_app.Visible = False
        except Exception:
            pass # Ignore if Visible property cannot be set
        try:
            excel_app.DisplayAlerts = False
        except Exception:
            pass
        try:
            excel_app.AskToUpdateLinks = False
        except Exception:
            pass
        
        print(f"[COM Extraction] Opening workbook: {excel_path}", file=sys.stderr)
        wb = excel_app.Workbooks.Open(excel_path, ReadOnly=True)
        
        sheets_data = []
        sheets_to_process = [wb.Worksheets(sheet_name)] if sheet_name else list(wb.Worksheets)
        
        for sheet_idx, ws in enumerate(sheets_to_process):
            try:
                sheet_name_actual = ws.Name
                print(f"[COM Extraction] Processing sheet: {sheet_name_actual}", file=sys.stderr)
                
                # Get used range to determine actual data bounds
                used_range = ws.UsedRange
                max_row = used_range.Rows.Count
                max_col = used_range.Columns.Count
                
                print(f"[COM Extraction] Sheet '{sheet_name_actual}' size: {max_row} rows x {max_col} cols", file=sys.stderr)
                
                # Extract all cell data with formatting (batch operations for speed)
                sheet_data = []
                merge_info = {}
                
                # Get all values and formulas at once (much faster than cell-by-cell)
                values = used_range.Value
                formulas = used_range.Formula
                
                # Convert to 2D list if single cell
                if max_row == 1 and max_col == 1:
                    values = [[values]]
                    formulas = [[formulas]]
                elif max_row == 1:
                    values = [list(values)]
                    formulas = [list(formulas)]
                elif max_col == 1:
                    values = [[v] for v in values]
                    formulas = [[f] for f in formulas]
                
                # Process data row by row with minimal formatting
                for row_idx in range(max_row):
                    row_data = []
                    for col_idx in range(max_col):
                        cell_value = values[row_idx][col_idx] if values and row_idx < len(values) and col_idx < len(values[row_idx]) else None
                        cell_formula = formulas[row_idx][col_idx] if formulas and row_idx < len(formulas) and col_idx < len(formulas[row_idx]) else None
                        
                        # Only create cell data if there's content
                        if cell_value is not None or (cell_formula and cell_formula != cell_value):
                            cell_data = {
                                'v': cell_value,
                                'm': str(cell_value) if cell_value is not None else '',
                            }
                            # Add formula if different from value
                            if cell_formula and str(cell_formula).startswith('='):
                                cell_data['f'] = cell_formula
                            
                            row_data.append(cell_data)
                        else:
                            row_data.append(None)
                    
                    sheet_data.append(row_data)
                
                print(f"[COM Extraction] ✓ Sheet '{sheet_name_actual}' extracted successfully", file=sys.stderr)
                
                # Build sheet object
                sheet_obj = {
                    'name': sheet_name_actual,
                    'data': sheet_data,
                    'config': {
                        'merge': merge_info,
                        'borderInfo': [],
                        'rowlen': {},
                        'columnlen': {}
                    },
                    'index': sheet_idx
                }
                
                sheets_data.append(sheet_obj)
                print(f"[COM Extraction] ✓ Extracted {len(sheet_data)} rows from '{sheet_name_actual}'", file=sys.stderr)
                
            except Exception as sheet_err:
                print(f"[COM Extraction] Error processing sheet: {sheet_err}", file=sys.stderr)
                continue
        
        # Close workbook and quit Excel
        wb.Close(SaveChanges=False)
        excel_app.Quit()
        pythoncom.CoUninitialize()
        
        print(f"[COM Extraction] ✓ Successfully extracted {len(sheets_data)} sheets", file=sys.stderr)
        return sheets_data
        
    except Exception as e:
        print(f"[COM Extraction] ERROR: {e}", file=sys.stderr)
        try:
            excel_app.Quit()
            pythoncom.CoUninitialize()
        except:
            pass
        return None


def _ensure_custom_xl_functions_registered():
    """Register Excel functions missing from xlcalculator (e.g., IFERROR)."""
    global _XL_CUSTOM_FUNCS_REGISTERED
    if _XL_CUSTOM_FUNCS_REGISTERED or not XL_CALC_AVAILABLE:
        return

    @xl.register()
    @xl.validate_args
    def IFERROR(
        value: func_xltypes.XlExpr,
        value_if_error: func_xltypes.XlExpr = lambda: func_xltypes.BLANK
    ):
        try:
            result = value()
            if isinstance(result, xlerrors.ExcelError):
                return value_if_error()
            return result
        except Exception:
            return value_if_error()

    _XL_CUSTOM_FUNCS_REGISTERED = True


def _coerce_model_constants_to_excel_types(model):
    """Wrap raw constants in ExcelType so xlcalculator functions can access .value."""
    if not XL_CALC_AVAILABLE:
        return
    for cell in model.cells.values():
        if cell.formula is not None:
            continue
        if cell.value is None:
            continue
        if isinstance(cell.value, func_xltypes.ExcelType):
            continue
        try:
            cell.value = func_xltypes.ExcelType.cast_from_native(cell.value)
        except Exception:
            # Leave value as-is if casting fails; evaluator will attempt later
            pass


def normalize_sheet_name(sheet_name: str) -> str:
    """Normalize sheet name by stripping whitespace and converting to lowercase."""
    return sheet_name.strip().lower()


def find_sheet_match(expected_sheet: str, available_sheets: List[str]) -> str:
    """
    Find the best matching sheet name from available sheets, ignoring case and spaces.
    
    Args:
        expected_sheet: The expected sheet name
        available_sheets: List of actual sheet names in the workbook
        
    Returns:
        The matching sheet name from available_sheets, or None if no match found
    """
    normalized_expected = normalize_sheet_name(expected_sheet)
    
    # First try exact match after normalization
    for sheet in available_sheets:
        if normalize_sheet_name(sheet) == normalized_expected:
            return sheet
    
    # If no exact match, try partial matches (in case of slight variations)
    for sheet in available_sheets:
        if normalized_expected in normalize_sheet_name(sheet) or normalize_sheet_name(sheet) in normalized_expected:
            return sheet
    
    return None


def generate_pdf_from_excel_sheet(excel_path: str, sheet_name: str, output_path: str) -> bool:
    """Generate PDF directly from Excel sheet using Excel COM automation to preserve all formatting."""
    try:
        print(f"[PDF Generator] Starting PDF generation for sheet: {sheet_name}", file=sys.stderr)
        print(f"[PDF Generator] COM_AVAILABLE: {COM_AVAILABLE}", file=sys.stderr)
        print(f"[PDF Generator] Input Excel: {excel_path}", file=sys.stderr)
        print(f"[PDF Generator] Output PDF: {output_path}", file=sys.stderr)
        
        if COM_AVAILABLE and USE_COM_INTERFACE:
            # Use Excel COM automation for exact formatting preservation
            print(f"[PDF Generator] Using Excel COM automation for exact formatting", file=sys.stderr)
            excel = None
            workbook = None
            co_initialized = False
            pythoncom = None
            try:
                import pythoncom as _pythoncom
                pythoncom = _pythoncom
                try:
                    pythoncom.CoInitialize()
                    co_initialized = True
                    print(f"[PDF Generator] CoInitialize called successfully", file=sys.stderr)
                except Exception as init_error:
                    print(f"[PDF Generator] Warning: Failed to CoInitialize COM: {init_error}", file=sys.stderr)
                
                print(f"[PDF Generator] Initializing Excel COM...", file=sys.stderr)
                excel = win32com.client.DispatchEx("Excel.Application")
                try:
                    excel.Visible = False
                except Exception as e:
                    print(f"[PDF Generator] Warning: Could not set Excel.Visible to False: {e}", file=sys.stderr)
                try:
                    excel.DisplayAlerts = False
                except Exception:
                    pass
                try:
                    excel.AskToUpdateLinks = False
                except Exception:
                    pass
                print(f"[PDF Generator] Excel COM initialized successfully", file=sys.stderr)
                
                # Open workbook
                print(f"[PDF Generator] Opening workbook: {os.path.abspath(excel_path)}", file=sys.stderr)
                workbook = excel.Workbooks.Open(os.path.abspath(excel_path), ReadOnly=True)
                print(f"[PDF Generator] Workbook opened, total sheets: {workbook.Sheets.Count}", file=sys.stderr)
                
                # Find and select the sheet (case-insensitive matching)
                sheet_found = False
                actual_sheet_name = None
                for sheet in workbook.Sheets:
                    print(f"[PDF Generator] Found sheet: {sheet.Name}", file=sys.stderr)
                    if normalize_sheet_name(sheet.Name) == normalize_sheet_name(sheet_name):
                        sheet.Select()
                        # Auto-fit columns to prevent ######## display for numeric values (skip for index sheet)
                        try:
                            if normalize_sheet_name(sheet.Name) != 'index':
                                sheet.Columns.AutoFit()
                        except Exception as e:
                            print(f"[PDF Generator] Warning: Could not AutoFit columns: {str(e)}", file=sys.stderr)
                        
                        sheet_found = True
                        actual_sheet_name = sheet.Name
                        print(f"[PDF Generator] Sheet '{actual_sheet_name}' selected (matched from '{sheet_name}')", file=sys.stderr)
                        
                        # Configure PageSetup for professional PDF output
                        print(f"[PDF Generator] Configuring PageSetup for '{actual_sheet_name}'...", file=sys.stderr)
                        try:
                            page_setup = sheet.PageSetup
                            page_setup.Zoom = False
                            page_setup.FitToPagesWide = 1
                            page_setup.FitToPagesTall = False  # Permit multi-page length
                            page_setup.CenterHorizontally = True
                            
                            # Smart orientation detection
                            if sheet.UsedRange.Columns.Count > 10:
                                page_setup.Orientation = 2  # xlLandscape
                            else:
                                page_setup.Orientation = 1  # xlPortrait
                                
                            # Professional margins
                            page_setup.LeftMargin = excel.InchesToPoints(0.25)
                            page_setup.RightMargin = excel.InchesToPoints(0.25)
                            page_setup.TopMargin = excel.InchesToPoints(0.4)
                            page_setup.BottomMargin = excel.InchesToPoints(0.4)
                            print(f"[PDF Generator] PageSetup configured successfully", file=sys.stderr)
                        except Exception as ps_error:
                            print(f"[PDF Generator] Warning: Could not configure PageSetup: {ps_error}", file=sys.stderr)
                            
                        break
                
                if not sheet_found:
                    print(f"[PDF Generator] ERROR: Sheet '{sheet_name}' not found in workbook (tried case-insensitive matching)", file=sys.stderr)
                    return False
                
                # Determine if we should ignore print areas for this specific export
                ignore_print_areas = True
                try:
                    norm_actual = normalize_sheet_name(actual_sheet_name)
                    if norm_actual in ['cover page', 'coverpage', 'cover', 'index'] and workbook.ActiveSheet.PageSetup.PrintArea:
                        ignore_print_areas = False
                        print(f"[PDF Generator] Respecting PrintArea for designed sheet: {actual_sheet_name}", file=sys.stderr)
                except:
                    pass

                # Export as PDF with optimal settings
                print(f"[PDF Generator] Exporting to PDF: {os.path.abspath(output_path)}", file=sys.stderr)
                workbook.ActiveSheet.ExportAsFixedFormat(
                    Type=0,  # xlTypePDF
                    Filename=os.path.abspath(output_path),
                    Quality=0,  # 0 = Standard quality (faster, smaller file)
                    IncludeDocProperties=True,
                    IgnorePrintAreas=ignore_print_areas,
                    OpenAfterPublish=False
                )
                
                print(f"[PDF Generator] PDF export completed", file=sys.stderr)
                print(f"[PDF Generator] PDF generated successfully using Excel COM: {output_path}", file=sys.stderr)
                return True
                
            except Exception as com_error:
                print(f"❌ [PDF Generator] Excel COM error: {str(com_error)}", file=sys.stderr)
                import traceback
                traceback.print_exc(file=sys.stderr)
                return False
            finally:
                if workbook is not None:
                    try:
                        workbook.Close(SaveChanges=False)
                    except Exception:
                        pass
                if excel:
                    try:
                        excel.Quit()
                    except Exception:
                        pass
                if co_initialized and pythoncom is not None:
                    try:
                        pythoncom.CoUninitialize()
                    except Exception:
                        pass
        else:
            # LibreOffice-based export for Linux/Render (preserves layout better than pandas/FPDF)
            print(f"⚠️ [PDF Generator] COM not available, using LibreOffice export", file=sys.stderr)
            return _generate_pdf_via_libreoffice(excel_path, sheet_name, output_path)
            
    except Exception as e:
        print(f"❌ [PDF Generator] Error generating PDF from Excel sheet: {str(e)}", file=sys.stderr)
        import traceback
        traceback.print_exc(file=sys.stderr)
        return False


def generate_pdf_fallback(excel_path: str, sheet_name: str, output_path: str) -> bool:
    """Fallback PDF generation using pandas (no formatting preservation)."""
    try:
        from fpdf import FPDF
        
        class SimplePDF(FPDF):
            def header(self):
                self.set_font('Arial', 'B', 12)
                self.cell(0, 8, f'{sheet_name} Sheet', 0, 1, 'C')
                self.ln(2)
            
            def footer(self):
                self.set_y(-15)
                self.set_font('Arial', 'I', 8)
                self.cell(0, 10, f'Page {self.page_no()}', 0, 0, 'C')
        
        # Read Excel sheet
        df = pd.read_excel(excel_path, sheet_name=sheet_name, engine='openpyxl', header=None)
        df = df.replace([pd.NA, np.inf, -np.inf], '')
        df = df.fillna('')

        # Create PDF
        pdf = SimplePDF()
        pdf.set_auto_page_break(auto=True, margin=15)
        pdf.add_page()
        pdf.set_font('Arial', '', 8)

        # Calculate column widths
        max_cols = len(df.columns)
        page_width = pdf.w - 30
        col_width = min(page_width / max_cols, 25)

        # Add data rows
        for row_idx in range(len(df)):
            row_data = df.iloc[row_idx]
            if pdf.get_y() > 250:
                pdf.add_page()
                pdf.set_font('Arial', '', 8)

            for col_idx in range(max_cols):
                value = str(row_data.iloc[col_idx]) if col_idx < len(row_data) else ''
                if len(value) > 15:
                    value = value[:12] + '...'

                if row_idx == 0:
                    pdf.set_fill_color(240, 240, 240)
                    pdf.cell(col_width, 8, value, 1, 0, 'C', 1)
                else:
                    pdf.cell(col_width, 8, value, 1, 0, 'C', 0)
            pdf.ln()

        pdf.output(output_path)
        print(f"PDF generated using fallback method: {output_path}", file=sys.stderr)
        return True

    except Exception as e:
        print(f"Fallback PDF generation failed: {str(e)}", file=sys.stderr)
        return False


def generate_pdfs_for_all_sheets(excel_path: str, output_dir: str, include_sheets: Optional[List[str]] = None, excluded_sheets: Optional[List[str]] = None, index_sheet: Optional[str] = None) -> Dict[str, Any]:
    """
    Generate individual PDF files for ALL sheets in the Excel workbook (excluding Assumptions sheet).
    Uses Excel COM automation to preserve formatting with better page fitting.
    
    Args:
        excel_path: Path to the Excel file
        output_dir: Directory to save the PDF files
        include_sheets: List of sheet names to include
        excluded_sheets: List of sheet names to exclude (overrides default)
        index_sheet: Specific Index sheet to include based on loan amount (e.g., 'Index < 20Lac')
        
    Returns:
        Dictionary with sheet names as keys and PDF file paths as values
    """
    print(f"\n{'='*80}", file=sys.stderr)
    print(f"📄 GENERATING PDFs FOR ALL EXCEL SHEETS", file=sys.stderr)
    print(f"{'='*80}\n", file=sys.stderr)

    # Sheets to exclude from PDF generation
    if excluded_sheets is not None:
        EXCLUDED_SHEETS = excluded_sheets
    else:
        EXCLUDED_SHEETS = [
            'Assumptions.1',  # Keep only this one which is typically internal
        ]
    
    pdf_files = {
        "sheets": {},
        "success_count": 0,
        "failed_count": 0,
        "total_sheets": 0,
        "excluded_sheets": [],
        "filtered_out_sheets": [],
        "requested_sheets": [],
        "sheet_status": [],
        "generated_files": []
    }

    include_filter = None
    sheet_status_summary = []
    requested_status_map = {}
    if include_sheets:
        include_filter = {
            _normalized_sheet_key(sheet): sheet.strip()
            for sheet in include_sheets
            if isinstance(sheet, str) and sheet.strip()
        }
        pdf_files["requested_sheets"] = list(include_filter.values())
        requested_status_map = {
            norm: {
                "sheet": original,
                "status": "pending",
                "reason": "Sheet not processed"
            }
            for norm, original in include_filter.items()
        }
    
    pythoncom = None
    co_initialized = False
    excel = None
    workbook = None

    # ---------------------------------------------------------
    # NON-COM IMPLEMENTATION (LibreOffice / Linux Support)
    # ---------------------------------------------------------
    if not USE_COM_INTERFACE:
        print(f"[Multi-PDF Generator] Using LibreOffice (No-COM) Mode", file=sys.stderr)
        import subprocess
        
        # Ensure output directory exists
        os.makedirs(output_dir, exist_ok=True)

        # We will loop through requested sheets (or all sheets) and generate PDFs
        # LibreOffice headless converts individual sheets by hiding others or printing ranges.
        # However, a simpler approach for 'specific sheets' is:
        # 1. Create a temporary copy of the workbook.
        # 2. Use openpyxl to delete all sheets EXCEPT the one we want.
        # 3. Convert that temporary workbook to PDF using soffice.
        
        temp_processing_dir = os.path.join(_get_safe_temp_dir(), "temp_libre_processing")
        os.makedirs(temp_processing_dir, exist_ok=True)

        try:
             # Load the workbook structure once to get sheet names
            wb_structure = load_workbook(excel_path, read_only=False, keep_vba=True)
            all_sheet_names = wb_structure.sheetnames
            try:
                wb_structure.close()
            except Exception:
                pass

            pdf_files["total_sheets"] = len(all_sheet_names)

            sheets_to_process = []
            matched_requested_norms = set()
            if include_sheets:
                # Filter requested
                 for sheet_name in all_sheet_names:
                    normalized_name = _normalized_sheet_key(sheet_name)
                    if normalized_name in include_filter:
                         sheets_to_process.append(sheet_name)
                         matched_requested_norms.add(normalized_name)
            else:
                # Use default excluded list if not provided
                safe_excluded = excluded_sheets if excluded_sheets else ['Assumptions.1']
                exclude_normalized = { _normalized_sheet_key(s) for s in safe_excluded }
                
                for sheet_name in all_sheet_names:
                     if _normalized_sheet_key(sheet_name) not in exclude_normalized:
                         sheets_to_process.append(sheet_name)
            
            print(f"[Multi-PDF Generator] Sheets to process (LibreOffice): {sheets_to_process}", file=sys.stderr)

            # Record missing requested sheets (do not fail; just report them)
            if include_filter:
                for norm_key, original_name in include_filter.items():
                    if norm_key not in matched_requested_norms and norm_key in requested_status_map:
                        requested_status_map[norm_key]["status"] = "missing"
                        requested_status_map[norm_key]["reason"] = "Sheet not found in workbook"

            for sheet_name in sheets_to_process:
                print(f"[Multi-PDF Generator] Processing sheet: {sheet_name}", file=sys.stderr)
                
                # Create a specific temp file for this sheet
                run_id = uuid.uuid4().hex[:8]
                temp_excel_name = f"temp_{run_id}_{_normalized_sheet_key(sheet_name)}.xlsx"
                temp_excel_path = os.path.join(temp_processing_dir, temp_excel_name)
                
                # Copy original to temp
                shutil.copy2(excel_path, temp_excel_path)
                
                # Open with OpenPyXL and delete other sheets
                try:
                    wb_temp = load_workbook(temp_excel_path)
                    for s in wb_temp.sheetnames:
                        if s != sheet_name:
                            del wb_temp[s]
                    try:
                        _apply_linux_page_setup(wb_temp[sheet_name], sheet_name)
                    except Exception as setup_error:
                        print(f"[Multi-PDF Generator] Warning: Could not apply page setup for '{sheet_name}': {setup_error}", file=sys.stderr)
                    wb_temp.save(temp_excel_path)
                    wb_temp.close()
                except Exception as e:
                     print(f"[Multi-PDF Generator] Error preparing temp excel for {sheet_name}: {e}", file=sys.stderr)
                     continue

                # Define output PDF path
                pdf_filename = f"{_normalized_sheet_key(sheet_name)}.pdf"
                pdf_output_path = os.path.join(output_dir, pdf_filename)
                
                # Run LibreOffice conversion
                # soffice --headless --convert-to pdf --outdir <output_dir> <input_file>
                cmd = [
                    "soffice",
                    "--headless",
                    "--convert-to", "pdf",
                    "--outdir", output_dir,
                    temp_excel_path
                ]
                
                print(f"[Multi-PDF Generator] running: {' '.join(cmd)}", file=sys.stderr)
                try:
                    result = subprocess.run(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, timeout=60)
                    if result.returncode == 0:
                        # LibreOffice output filename matches input filename but .pdf
                        # e.g. temp_123_sheetname.pdf
                        generated_temp_pdf_name = temp_excel_name.replace(".xlsx", ".pdf")
                        generated_temp_pdf_path = os.path.join(output_dir, generated_temp_pdf_name)
                        
                        if os.path.exists(generated_temp_pdf_path):
                            # Rename to desired final name
                            if os.path.exists(pdf_output_path):
                                os.remove(pdf_output_path)
                            shutil.move(generated_temp_pdf_path, pdf_output_path)

                            file_size = 0
                            try:
                                file_size = os.path.getsize(pdf_output_path)
                            except Exception:
                                file_size = 0
                            
                            pdf_files["generated_files"].append(pdf_output_path)
                            pdf_files["success_count"] += 1
                            pdf_files["sheets"][sheet_name] = {
                                "pdf_path": pdf_output_path,
                                "pdf_filename": os.path.basename(pdf_output_path),
                                "file_size": file_size,
                                "status": "success"
                            }
                            sheet_status_summary.append({
                                "sheet": sheet_name,
                                "status": "success",
                                "reason": "Generated via LibreOffice",
                                "path": pdf_output_path
                            })
                            
                            # Update status
                            norm = _normalized_sheet_key(sheet_name)
                            if norm in requested_status_map:
                                requested_status_map[norm]["status"] = "success"
                                requested_status_map[norm]["path"] = pdf_output_path
                                requested_status_map[norm]["reason"] = "Generated via LibreOffice"
                        else:
                             print(f"[Multi-PDF Generator] Error: PDF not found after generation: {generated_temp_pdf_path}", file=sys.stderr)
                             pdf_files["failed_count"] += 1
                             pdf_files["sheets"][sheet_name] = {"status": "failed", "error": "PDF not created by LibreOffice"}
                             sheet_status_summary.append({
                                "sheet": sheet_name,
                                "status": "failed",
                                "reason": "PDF not created by LibreOffice"
                             })
                             norm = _normalized_sheet_key(sheet_name)
                             if norm in requested_status_map:
                                requested_status_map[norm]["status"] = "failed"
                                requested_status_map[norm]["reason"] = "PDF not created by LibreOffice"
                    else:
                        err_text = result.stderr.decode(errors="replace")
                        print(f"[Multi-PDF Generator] LibreOffice Error: {err_text}", file=sys.stderr)
                        pdf_files["failed_count"] += 1
                        pdf_files["sheets"][sheet_name] = {"status": "failed", "error": err_text}
                        sheet_status_summary.append({
                            "sheet": sheet_name,
                            "status": "failed",
                            "reason": "LibreOffice conversion failed",
                            "error": err_text
                        })
                        norm = _normalized_sheet_key(sheet_name)
                        if norm in requested_status_map:
                            requested_status_map[norm]["status"] = "failed"
                            requested_status_map[norm]["reason"] = "LibreOffice conversion failed"
                        
                except Exception as e:
                    print(f"[Multi-PDF Generator] Subprocess Error: {e}", file=sys.stderr)
                    pdf_files["failed_count"] += 1
                    pdf_files["sheets"][sheet_name] = {"status": "failed", "error": str(e)}
                    sheet_status_summary.append({
                        "sheet": sheet_name,
                        "status": "failed",
                        "reason": "LibreOffice subprocess error",
                        "error": str(e)
                    })
                    norm = _normalized_sheet_key(sheet_name)
                    if norm in requested_status_map:
                        requested_status_map[norm]["status"] = "failed"
                        requested_status_map[norm]["reason"] = "LibreOffice subprocess error"

                # Cleanup temp excel
                try:
                    os.remove(temp_excel_path)
                except:
                    pass

        except Exception as e:
             print(f"[Multi-PDF Generator] Global Error (LibreOffice Loop): {e}", file=sys.stderr)
        
        finally:
            # Cleanup temp dir
            try:
                shutil.rmtree(temp_processing_dir)
            except:
                pass

        # Attach status summaries
        if include_filter:
            pdf_files["sheet_status"] = [requested_status_map[norm] for norm in include_filter.keys() if norm in requested_status_map]
        else:
            pdf_files["sheet_status"] = sheet_status_summary
                
        return pdf_files

    # ---------------------------------------------------------
    # COM IMPLEMENTATION (Windows Only)
    # ---------------------------------------------------------
    try:
        if not COM_AVAILABLE:
            print(f"❌ Excel COM not available. Cannot generate PDFs with formatting.", file=sys.stderr)
            return pdf_files
        
        try:
            import pythoncom as _pythoncom
            pythoncom = _pythoncom
            pythoncom.CoInitialize()
            co_initialized = True
            print(f"[Multi-PDF Generator] CoInitialize called successfully", file=sys.stderr)
        except Exception as init_error:
            print(f"[Multi-PDF Generator] ERROR: Could not initialize COM: {init_error}", file=sys.stderr)
            raise

        # Ensure output directory exists
        os.makedirs(output_dir, exist_ok=True)
        
        # Open workbook with Excel COM
        print(f"[Multi-PDF Generator] Opening workbook: {excel_path}", file=sys.stderr)
        print(f"[Multi-PDF Generator] Using Excel COM Automation (DispatchEx)", file=sys.stderr)
        excel = win32com.client.DispatchEx("Excel.Application")
        try:
            excel.Visible = False
            excel.ScreenUpdating = False
        except Exception as e:
            print(f"[Multi-PDF Generator] Warning: Could not set Excel.Visible/ScreenUpdating to False: {e}", file=sys.stderr)
        try:
            excel.DisplayAlerts = False
        except Exception:
            pass
        try:
            excel.AskToUpdateLinks = False
        except Exception:
            pass
        
        workbook = excel.Workbooks.Open(os.path.abspath(excel_path), ReadOnly=True)
        total_sheets = workbook.Sheets.Count
        pdf_files["total_sheets"] = total_sheets
        
        print(f"[Multi-PDF Generator] Found {total_sheets} sheets", file=sys.stderr)
        print(f"{'─'*80}\n", file=sys.stderr)
        
        # Generate PDF for each sheet
        for sheet_idx in range(1, total_sheets + 1):
            sheet = workbook.Sheets(sheet_idx)
            sheet_name = sheet.Name
            
            normalized_sheet = sheet_name.strip()
            normalized_key = re.sub(r'[\s_\-]+', '', normalized_sheet.lower())

            # Skip excluded sheets (like Assumptions)
            if sheet_name in EXCLUDED_SHEETS:
                print(f"[{sheet_idx}/{total_sheets}] ⏭️  Skipping sheet: '{sheet_name}' (excluded)", file=sys.stderr)
                pdf_files["excluded_sheets"].append(sheet_name)
                if include_filter and normalized_key in requested_status_map:
                    requested_status_map[normalized_key]["status"] = "failed"
                    requested_status_map[normalized_key]["reason"] = "Sheet excluded from PDF generation"
                else:
                    sheet_status_summary.append({
                        "sheet": sheet_name,
                        "status": "excluded",
                        "reason": "Sheet excluded from PDF generation"
                    })
                continue

            # Handle Index sheet filtering based on loan amount condition
            is_index_sheet = sheet_name.strip().startswith('Index')
            if is_index_sheet and index_sheet:
                # If a specific index sheet is specified, only include that one
                if sheet_name.strip() != index_sheet.strip():
                    print(f"[{sheet_idx}/{total_sheets}] ⏭️  Skipping sheet: '{sheet_name}' (using '{index_sheet}' instead)", file=sys.stderr)
                    pdf_files["filtered_out_sheets"].append(sheet_name)
                    sheet_status_summary.append({
                        "sheet": sheet_name,
                        "status": "filtered",
                        "reason": f"Using Index sheet: {index_sheet}"
                    })
                    continue
                else:
                    print(f"[{sheet_idx}/{total_sheets}] Including Index sheet: '{sheet_name}' (loan amount condition match)", file=sys.stderr)
            elif is_index_sheet and not index_sheet:
                # If no specific index sheet is specified, skip all index sheets
                print(f"[{sheet_idx}/{total_sheets}] ⏭️  Skipping sheet: '{sheet_name}' (no index sheet specified)", file=sys.stderr)
                pdf_files["filtered_out_sheets"].append(sheet_name)
                sheet_status_summary.append({
                    "sheet": sheet_name,
                    "status": "filtered",
                    "reason": "Index sheet not applicable for this report"
                })
                continue

            if include_filter and normalized_key not in include_filter:
                print(f"[{sheet_idx}/{total_sheets}] ⏭️  Skipping sheet: '{sheet_name}' (not in requested list)", file=sys.stderr)
                pdf_files["filtered_out_sheets"].append(sheet_name)
                sheet_status_summary.append({
                    "sheet": sheet_name,
                    "status": "filtered",
                    "reason": "Sheet not included in requested list"
                })
                continue
            
            print(f"[{sheet_idx}/{total_sheets}] Processing sheet: '{sheet_name}'", file=sys.stderr)
            
            # Create PDF filename (sanitize sheet name)
            safe_sheet_name = re.sub(r'[<>:"/\\|?*]', '_', sheet_name)
            pdf_filename = f"sheet_{sheet_idx}_{safe_sheet_name}.pdf"
            pdf_path = os.path.join(output_dir, pdf_filename)
            
            try:
                # Select the sheet
                sheet.Select()
                
                # Get normalized sheet name for special handling
                normalized_sheet = normalize_sheet_name(sheet_name)
                
                # Special sheets that need to fill the page (Cover, Index, etc.)
                special_fill_sheets = ['cover page', 'coverpage', 'cover', 'index', 'summary', 'report summary', 'introduction']
                is_designed_sheet = any(s in normalized_sheet for s in special_fill_sheets)
                
                # Auto-fit columns to prevent ######## display for numeric values (skip for designed sheets)
                # Designed sheets like Cover Page/Index usually shouldn't be auto-fitted as it ruins layout
                try:
                    if not is_designed_sheet:
                        sheet.Columns.AutoFit()
                except Exception as e:
                    print(f"   ⚠️  Warning: Could not AutoFit columns on '{sheet_name}': {str(e)}", file=sys.stderr)
                
                # CRITICAL: Define PageSetup object immediately
                page_setup = sheet.PageSetup

                # Hide consecutive empty rows (skip for designed sheets to preserve manual layout)
                try:
                    # Logic to hide empty rows (SKIP FOR DESIGNED SHEETS)
                    if not is_designed_sheet and include_filter is None:  
                        # Determine the used range manually to be safe
                        used_range = sheet.UsedRange
                        first_row = used_range.Row
                        total_rows = used_range.Rows.Count
                        last_row = first_row + total_rows - 1
                        
                        first_col = used_range.Column
                        last_col = first_col + used_range.Columns.Count - 1
                        
                        # Initialize rows_to_hide
                        rows_to_hide = []
                        consecutive_empty_start = None
                        consecutive_empty_count = 0
                        
                        for row_idx in range(first_row, last_row + 1):
                            # Check if entire row is empty
                            row_is_empty = True
                            for col_idx in range(first_col, last_col + 1):
                                cell_value = sheet.Cells(row_idx, col_idx).Value
                                if cell_value is not None and str(cell_value).strip() != '':
                                    row_is_empty = False
                                    break
                            
                            if row_is_empty:
                                if consecutive_empty_start is None:
                                    consecutive_empty_start = row_idx
                                consecutive_empty_count += 1
                            else:
                                # If we had more than 1 consecutive empty rows, mark them for hiding (keep 1 visible)
                                if consecutive_empty_count > 1:
                                    # Keep the first empty row visible, hide the rest
                                    for hide_row in range(consecutive_empty_start + 1, consecutive_empty_start + consecutive_empty_count):
                                        rows_to_hide.append(hide_row)
                                consecutive_empty_start = None
                                consecutive_empty_count = 0
                        
                        # Handle trailing empty rows (hide all if more than 1)
                        if consecutive_empty_count > 1:
                            for hide_row in range(consecutive_empty_start + 1, consecutive_empty_start + consecutive_empty_count):
                                rows_to_hide.append(hide_row)
                        
                        # Hide the rows
                        rows_hidden = 0
                        for row_idx in rows_to_hide:
                            try:
                                sheet.Rows(row_idx).Hidden = True
                                rows_hidden += 1
                            except:
                                pass
                        
                        if rows_hidden > 0:
                            print(f"   👁️  Hidden {rows_hidden} empty rows in '{sheet_name}'", file=sys.stderr)
                            
                except Exception as e:
                    print(f"   ⚠️  Could not hide empty rows: {str(e)}", file=sys.stderr)
                
                # Sheets that typically need landscape due to many columns
                wide_sheets = ['workings for pl & bs', 'workings for plbs', 'workings for pl bs 2', 
                              'plbs', 'ratio', 'workings', 'working for plbs', 'working for plbs1', 
                              'working for plbs2', 'workings for pl bs', 'tl workings']
                
                # Default behavior for print areas: Respect them to exclude side instructions
                ignore_print_areas = False
                
                if is_designed_sheet:
                    # NATIVE EXPORT with SCALING OVERRIDES
                    # 1. Use Native PrintArea (ignore_print_areas=False) to respect "White Page" area and exclude side notes.
                    # 2. Force A4 + Fit-to-Page to ensure that specific PrintArea fills the PDF accurately.
                    
                    print(f"   📄 Cover '{sheet_name}': Using Native PrintArea + Force A4 Scaling", file=sys.stderr)
                    
                    # Do NOT clear PrintArea. Trust the template's border definition.
                    # page_setup.PrintArea = "" 
                    
                    page_setup.PaperSize = 9 # A4
                    page_setup.Zoom = False
                    page_setup.FitToPagesWide = 1
                    page_setup.FitToPagesTall = 1 
                    page_setup.CenterHorizontally = True
                    
                    ignore_print_areas = False 
                
                else:
                    # Configure page setup for better fitting using the specific sheet object
                    # page_setup was already defined above
                    page_setup.Zoom = False  # Disable fixed zoom (required for FitToPages)
                    page_setup.PaperSize = 9  # A4 size
                    
                    # Narrow Margins to allow more rows
                    # Standard is ~0.75 inch. We set to 0.25 inch (approx 18 points)
                    page_setup.HeaderMargin = excel.InchesToPoints(0.2)
                    page_setup.FooterMargin = excel.InchesToPoints(0.2)
                    page_setup.LeftMargin = excel.InchesToPoints(0.25)
                    page_setup.RightMargin = excel.InchesToPoints(0.25)
                    page_setup.TopMargin = excel.InchesToPoints(0.3)
                    page_setup.BottomMargin = excel.InchesToPoints(0.3)

                    # Dynamic Fit Logic: Target ~60 rows per page
                    # Calculate required pages based on row count
                    fit_tall = False
                    try:
                        used_rows = sheet.UsedRange.Rows.Count
                        # Formula: Total Rows / 60 -> Rounded UP.
                        # e.g. 120 -> 2. 61 -> 2. 
                        # Note: This balances the pages (e.g. 90 -> 2 pages @ 45 rows).
                        # To force exactly 60 on P1, we'd need manual page breaks, which is complex.
                        # This scaling ensures reasonable density.
                        pages_tall = (used_rows + 59) // 60
                        if pages_tall < 1: pages_tall = 1
                        fit_tall = pages_tall
                    except:
                        fit_tall = False

                    if normalized_sheet in wide_sheets or any(ws in normalized_sheet for ws in wide_sheets):
                        # Wide data sheets: Landscape, fit to width
                        page_setup.FitToPagesWide = 1
                        page_setup.FitToPagesTall = fit_tall 
                        page_setup.Orientation = 2  # xlLandscape
                        page_setup.CenterHorizontally = True
                        print(f"   📄 Wide sheet '{sheet_name}': Landscape, Fit target {fit_tall} pages", file=sys.stderr)
                    else:
                        # Regular sheets: Portrait, fit to width
                        page_setup.FitToPagesWide = 1
                        page_setup.FitToPagesTall = fit_tall 
                        page_setup.CenterHorizontally = True
                        
                        # Check if content is too wide by examining used range
                        try:
                            used_range = sheet.UsedRange
                            last_col = used_range.Columns.Count
                            if last_col > 10:
                                page_setup.Orientation = 2  # xlLandscape
                                print(f"   📄 Sheet '{sheet_name}': Landscape (detected {last_col} columns)", file=sys.stderr)
                            else:
                                page_setup.Orientation = 1  # xlPortrait
                                print(f"   📄 Sheet '{sheet_name}': Portrait ({last_col} columns)", file=sys.stderr)
                        except:
                            page_setup.Orientation = 1  # Default to Portrait
                            print(f"   📄 Sheet '{sheet_name}': Portrait (default)", file=sys.stderr)

                # Export as PDF using the specific SHEET object instead of workbook
                sheet.ExportAsFixedFormat(
                    Type=0,  # xlTypePDF
                    Filename=os.path.abspath(pdf_path),
                    Quality=0,  # Standard quality
                    IncludeDocProperties=True,
                    IgnorePrintAreas=ignore_print_areas,
                    OpenAfterPublish=False
                )
                
                # Check if PDF was created
                if os.path.exists(pdf_path):
                    file_size = os.path.getsize(pdf_path)
                    print(f"   ✅ PDF created: {pdf_filename} ({file_size:,} bytes)", file=sys.stderr)
                    
                    pdf_files["sheets"][sheet_name] = {
                        "pdf_path": pdf_path,
                        "pdf_filename": pdf_filename,
                        "sheet_index": sheet_idx,
                        "file_size": file_size,
                        "status": "success"
                    }
                    pdf_files["success_count"] += 1
                    if include_filter:
                        if normalized_key in requested_status_map:
                            requested_status_map[normalized_key]["status"] = "success"
                            requested_status_map[normalized_key]["reason"] = "PDF generated successfully"
                    else:
                        sheet_status_summary.append({
                            "sheet": sheet_name,
                            "status": "success",
                            "reason": "PDF generated successfully"
                        })
                else:
                    print(f"   ❌ PDF file not created", file=sys.stderr)
                    pdf_files["sheets"][sheet_name] = {
                        "status": "failed",
                        "error": "PDF file not created"
                    }
                    pdf_files["failed_count"] += 1
                    failure_reason = "PDF file not created"
                    if include_filter:
                        if normalized_key in requested_status_map:
                            requested_status_map[normalized_key]["status"] = "failed"
                            requested_status_map[normalized_key]["reason"] = failure_reason
                    else:
                        sheet_status_summary.append({
                            "sheet": sheet_name,
                            "status": "failed",
                            "reason": failure_reason
                        })
                    
            except Exception as sheet_error:
                print(f"   ❌ Error generating PDF for sheet '{sheet_name}': {str(sheet_error)}", file=sys.stderr)
                pdf_files["sheets"][sheet_name] = {
                    "status": "failed",
                    "error": str(sheet_error)
                }
                pdf_files["failed_count"] += 1
                failure_reason = str(sheet_error)
                if include_filter:
                    if normalized_key in requested_status_map:
                        requested_status_map[normalized_key]["status"] = "failed"
                        requested_status_map[normalized_key]["reason"] = failure_reason
                else:
                    sheet_status_summary.append({
                        "sheet": sheet_name,
                        "status": "failed",
                        "reason": failure_reason
                    })
        
        # Close workbook and Excel
        workbook.Close(SaveChanges=False)
        workbook = None
        excel.Quit()
        excel = None
        
        print(f"\n{'─'*80}", file=sys.stderr)
        print(f"✅ PDF Generation Complete", file=sys.stderr)
        print(f"   Total Sheets: {pdf_files['total_sheets']}", file=sys.stderr)
        if pdf_files['requested_sheets']:
            print(f"   Requested: {len(pdf_files['requested_sheets'])} ({', '.join(pdf_files['requested_sheets'])})", file=sys.stderr)
        print(f"   Excluded: {len(pdf_files['excluded_sheets'])} ({', '.join(pdf_files['excluded_sheets']) if pdf_files['excluded_sheets'] else 'none'})", file=sys.stderr)
        if pdf_files['filtered_out_sheets']:
            print(f"   Filtered Out: {len(pdf_files['filtered_out_sheets'])} ({', '.join(pdf_files['filtered_out_sheets'])})", file=sys.stderr)
        print(f"   Successful: {pdf_files['success_count']}", file=sys.stderr)
        print(f"   Failed: {pdf_files['failed_count']}", file=sys.stderr)
        print(f"{'='*80}\n", file=sys.stderr)
        
        if include_filter:
            for norm_key, status_entry in requested_status_map.items():
                if status_entry["status"] == "pending":
                    status_entry["status"] = "failed"
                    status_entry["reason"] = "Requested sheet not found in workbook"
            pdf_files["sheet_status"] = list(requested_status_map.values())
        else:
            pdf_files["sheet_status"] = sheet_status_summary

        return pdf_files
        
    except Exception as e:
        print(f"❌ Error in multi-sheet PDF generation: {str(e)}", file=sys.stderr)
        import traceback
        traceback.print_exc(file=sys.stderr)
        return pdf_files
    finally:
        if workbook is not None:
            try:
                workbook.Close(SaveChanges=False)
            except Exception:
                pass
        if excel is not None:
            try:
                excel.Quit()
            except Exception:
                pass
        if co_initialized and pythoncom is not None:
            try:
                pythoncom.CoUninitialize()
                print(f"[Multi-PDF Generator] CoUninitialize completed", file=sys.stderr)
            except Exception as cleanup_error:
                print(f"[Multi-PDF Generator] Warning: CoUninitialize failed: {cleanup_error}", file=sys.stderr)

def generate_html_from_excel_com(excel_path: str, sheet_name: str, header_data: dict = None) -> tuple:
    """
    Generate HTML from Excel using COM automation to get calculated values.
    This ensures formulas are evaluated and we get the actual values.
    Returns: (html_content, json_data)
    """
    try:
        print(f"[HTML COM Generator] Starting HTML generation using Excel COM", file=sys.stderr)
        
        excel = win32com.client.DispatchEx("Excel.Application")
        try:
            excel.Visible = False
        except Exception as e:
            print(f"[HTML COM Generator] Warning: Could not set Excel.Visible to False: {e}", file=sys.stderr)
        try:
            excel.DisplayAlerts = False
        except Exception:
            pass
        try:
            excel.AskToUpdateLinks = False
        except Exception:
            pass
        
        wb = excel.Workbooks.Open(os.path.abspath(excel_path), ReadOnly=True)
        
        # Get all sheet names
        available_sheets = [ws.Name for ws in wb.Worksheets]
        
        # Find the matching sheet name (handles case and space differences)
        actual_sheet_name = find_sheet_match(sheet_name, available_sheets)
        if not actual_sheet_name:
            print(f"[HTML COM Generator] Sheet '{sheet_name}' not found (tried case-insensitive matching)", file=sys.stderr)
            print(f"[HTML COM Generator] Available sheets: {available_sheets}", file=sys.stderr)
            print(f"[HTML COM Generator] Template may use different sheet naming convention", file=sys.stderr)
            wb.Close(False)
            excel.Quit()
            return "", {}
        
        # Find the sheet
        sheet = None
        for ws in wb.Worksheets:
            if ws.Name == actual_sheet_name:
                sheet = ws
                break
        
        print(f"[HTML COM Generator] Processing sheet: {actual_sheet_name} (matched from '{sheet_name}')", file=sys.stderr)
        
        # Get used range
        used_range = sheet.UsedRange
        max_row = used_range.Rows.Count
        max_col = used_range.Columns.Count
        
        print(f"[HTML COM Generator] Processing {max_row} rows x {max_col} columns", file=sys.stderr)
        
        # Extract JSON data structure
        json_data = {
            "sheetName": actual_sheet_name,
            "data": {},
            "timestamp": datetime.datetime.now().isoformat()
        }
        
        # Extract firm details from the data for receipt header
        firm_name = ""
        proprietor = ""
        sector = ""
        nature_of_business = ""
        
        # Use provided header data if available
        if header_data:
            proprietor = header_data.get('proprietor', '')
            sector = header_data.get('sector', '')
            nature_of_business = header_data.get('natureOfBusiness', '')
        
        # Try to get firm details from common cell positions (only if not already provided)
        try:
            if max_row >= 3:
                firm_name_cell = sheet.Cells(3, 2).Value
                if firm_name_cell:
                    firm_name = str(firm_name_cell)
            
            if not proprietor and max_row >= 4:
                proprietor_cell = sheet.Cells(4, 2).Value
                if proprietor_cell:
                    proprietor = str(proprietor_cell)
            
            if not sector and max_row >= 6:
                sector_cell = sheet.Cells(6, 2).Value
                if sector_cell:
                    sector = str(sector_cell)
            
            if not nature_of_business and max_row >= 7:
                nature_cell = sheet.Cells(7, 2).Value
                if nature_cell:
                    nature_of_business = str(nature_cell)
        except:
            pass
        
        # Build HTML with modern professional styling
        html_parts = [
            "<!DOCTYPE html>",
            "<html lang='en'>",
            "<head>",
            "<meta charset='UTF-8'>",
            "<meta name='viewport' content='width=device-width, initial-scale=1.0'>",
            f"<title>Financial Report - {firm_name or sheet_name}</title>",
            "<link href='https://fonts.googleapis.com/css2?family=Manrope:wght@400;500;600;700;800&family=Inter:wght@300;400;500;600;700&display=swap' rel='stylesheet'>",
            "<link rel='stylesheet' href='https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.0.0-beta3/css/all.min.css'>", # For professional icons
            "<style>",
            "  :root {",
            "    --primary-purple: #7c3aed;",
            "    --primary-dark-purple: #6d28d9;",
            "    --primary-black: #1f2937;",
            "    --primary-light-black: #374151;",
            "    --ghost-white: #F8F8FF;",
            "    --success-green: #10b981;",
            "    --text-primary: var(--primary-black);",
            "    --text-secondary: #6b7280;",
            "    --bg-primary: #ffffff;",
            "    --bg-secondary: var(--ghost-white);",
            "    --bg-accent: #e5e7eb;",
            "    --border-color: #e5e7eb;",
            "    --shadow-soft: 0 2px 15px -3px rgba(0, 0, 0, 0.07), 0 10px 20px -2px rgba(0, 0, 0, 0.04);",
            "    --shadow-md: 0 4px 6px -1px rgba(0, 0, 0, 0.1);",
            "    --shadow-lg: 0 10px 15px -3px rgba(0, 0, 0, 0.1);",
            "    --shadow-xl: 0 20px 25px -5px rgba(0, 0, 0, 0.1);",
            "  }",
            "  ",
            "  * {",
            "    margin: 0;",
            "    padding: 0;",
            "    box-sizing: border-box;",
            "  }",
            "  ",
            "  body {",
            "    font-family: 'Inter', -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;",
            "    background-color: var(--bg-secondary);", # Ghost White background
            "    min-height: 100vh;",
            "    padding: 20px;",
            "    line-height: 1.6;",
            "    color: var(--text-primary);",
            "    -webkit-font-smoothing: antialiased;",
            "    -moz-osx-font-smoothing: grayscale;",
            "  }",
            "  ",
            "  .container {",
            "    max-width: 1400px;",
            "    margin: 0 auto;",
            "  }",
            "  ",
            "  .report-card {",
            "    background: var(--bg-primary);",
            "    border-radius: 12px;", # Rounded corners
            "    box-shadow: var(--shadow-soft);", # Soft shadow
            "    overflow: hidden;",
            "    animation: slideUp 0.6s ease-out;",
            "  }",
            "  ",
            "  @keyframes slideUp {",
            "    from {",
            "      opacity: 0;",
            "      transform: translateY(30px);",
            "    }",
            "    to {",
            "      opacity: 1;",
            "      transform: translateY(0);",
            "    }",
            "  }",
            "  ",
            "  /* Header Section */",
            "  .report-header {"
            "    padding: 32px 24px;",
            "    color: black;",
            "    position: relative;",
            "    overflow: hidden;",
            "    border-bottom: 1px solid rgba(255, 255, 255, 0.1);",
            "  }",
            "  ",
            "  .report-header::before {",
            "    content: '';",
            "    position: absolute;",
            "    top: 0;",
            "    right: 0;",
            "    width: 200px;",
            "    height: 200px;",
            "    background: radial-gradient(circle, rgba(255,255,255,0.15) 0%, transparent 70%);",
            "    border-radius: 50%;",
            "    transform: translate(30%, -30%);",
            "  }",
            "  ",
            "  .header-content {",
            "    position: relative;",
            "    z-index: 1;",
            "  }",
            "  ",
            "  .report-badge {",
            "    display: inline-flex;", # Use flex for icon alignment
            "    align-items: center;",
            "    gap: 8px;",
            "    background: rgba(255, 255, 255, 0.2);",
            "    backdrop-filter: blur(5px);",
            "    padding: 6px 16px;",
            "    border-radius: 50px;",
            "    font-size: 12px;",
            "    font-weight: 600;",
            "    letter-spacing: 0.5px;",
            "    text-transform: uppercase;",
            "    margin-bottom: 16px;",
            "  }",
            "  .report-badge i {",
            "    font-size: 14px;",
            "  }",
            "  ",
            "  .firm-name {",
            "    font-family: 'Manrope', sans-serif;", # Manrope for headings
            "    font-size: 28px;",
            "    font-weight: 700;",
            "    margin-bottom: 12px;",
            "    letter-spacing: -0.5px;",
            "  }",
            "  ",
            "  .firm-meta {",
            "    display: flex;",
            "    flex-wrap: wrap;",
            "    gap: 24px;",
            "    margin-top: 20px;",
            "    padding-top: 20px;",
            "    border-top: 1px solid rgba(255, 255, 255, 0.15);",
            "  }",
            "  ",
            "  .meta-item {",
            "    display: flex;",
            "    flex-direction: column;",
            "    gap: 4px;",
            "  }",
            "  ",
            "  .meta-label {",
            "    font-size: 11px;",
            "    font-weight: 500;",
            "    opacity: 0.9;",
            "    text-transform: uppercase;",
            "    letter-spacing: 0.8px;",
            "  }",
            "  ",
            "  .meta-value {",
            "    font-size: 15px;",
            "    font-weight: 600;",
            "  }",
            "  ",
            "  /* Stats Cards */",
            "  .stats-grid {",
            "    display: grid;",
            "    grid-template-columns: repeat(auto-fit, minmax(180px, 1fr));",
            "    gap: 1px;",
            "    background: var(--border-color);",
            "    border-bottom: 1px solid var(--border-color);",
            "  }",
            "  ",
            "  .stat-card {",
            "    background: var(--bg-primary);",
            "    padding: 24px 20px;",
            "    text-align: center;",
            "    transition: all 0.3s ease;",
            "  }",
            "  ",
            "  .stat-card:hover {",
            "    background: var(--bg-secondary);",
            "    transform: translateY(-2px);",
            "  }",
            "  ",
            "  .stat-icon {",
            "    width: 40px;",
            "    height: 40px;",
            "    margin: 0 auto 12px;",
            "    border-radius: 8px;",
            "    display: flex;",
            "    align-items: center;",
            "    justify-content: center;",
            "    font-size: 18px;",
            "    color: black;",
            "  }",
            "  ",
            "  .stat-label {",
            "    font-size: 11px;",
            "    font-weight: 600;",
            "    color: var(--text-secondary);",
            "    text-transform: uppercase;",
            "    letter-spacing: 0.8px;",
            "    margin-bottom: 6px;",
            "  }",
            "  ",
            "  .stat-value {",
            "    font-size: 16px;",
            "    font-weight: 700;",
            "    color: var(--text-primary);",
            "    font-family: 'Manrope', sans-serif;", # Manrope for values
            "  }",
            "  ",
            "  /* Table Section */",
            "  .table-section {",
            "    padding: 32px 24px;",
            "  }",
            "  ",
            "  .section-title {",
            "    font-family: 'Manrope', sans-serif;", # Manrope for titles
            "    font-size: 20px;",
            "    font-weight: 700;",
            "    color: var(--text-primary);",
            "    margin-bottom: 6px;",
            "  }",
            "  ",
            "  .section-subtitle {",
            "    font-size: 13px;",
            "    color: var(--text-secondary);",
            "    margin-bottom: 24px;",
            "  }",
            "  ",
            "  .table-wrapper {",
            "    overflow-x: auto;",
            "    border-radius: 8px;",
            "    border: 1px solid var(--border-color);",
            "  }",
            "  ",
            "  table {",
            "    width: 100%;",
            "    border-collapse: collapse;",
            "    background: var(--bg-primary);",
            "  }",
            "  ",
            "  thead {",
            "    background: var(--bg-accent);",
            "    position: sticky;",
            "    top: 0;",
            "    z-index: 10;",
            "  }",
            "  ",
            "  thead th {",
            "    padding: 14px 20px;",
            "    text-align: left;",
            "    font-size: 11px;",
            "    font-weight: 700;",
            "    color: var(--text-primary);",
            "    text-transform: uppercase;",
            "    letter-spacing: 0.8px;",
            "    border-bottom: 1px solid var(--border-color);",
            "  }",
            "  ",
            "  thead th:last-child {",
            "    text-align: right;",
            "  }",
            "  ",
            "  tbody tr {",
            "    border-bottom: 1px solid var(--border-color);",
            "    transition: all 0.2s ease;",
            "  }",
            "  ",
            "  tbody tr:hover {",
            "    background: var(--bg-secondary);",
            "  }",
            "  ",
            "  tbody tr:last-child {",
            "    border-bottom: none;",
            "  }",
            "  ",
            "  td {",
            "    padding: 16px 20px;",
            "    font-size: 13px;",
            "    color: var(--text-primary);",
            "  }",
            "  ",
            "  .item-name {",
            "    font-weight: 500;",
            "  }",
            "  ",
            "  .item-value {",
            "    text-align: right;",
            "    font-weight: 600;",
            "    font-family: 'Inter', monospace;",
            "    font-size: 14px;",
            "  }",
            "  ",
            "  .currency::before {",
            "    content: '₹ ';",
            "    color: var(--text-secondary);",
            "    margin-right: 2px;",
            "    font-weight: 500;",
            "  }",
            "  ",
            "  /* Special Row Styles */",
            "  .section-header {",
            "    background: var(--bg-accent) !important;",
            "  }",
            "  ",
            "  .section-header td {",
            "    font-family: 'Manrope', sans-serif !important;",
            "    font-weight: 700 !important;",
            "    font-size: 13px !important;",
            "    color: var(--text-primary) !important;",
            "    padding: 12px 20px !important;",
            "    text-transform: uppercase;",
            "    letter-spacing: 0.5px;",
            "  }",
            "  ",
            "  .total-row {",
            "  }",
            "  ",
            "  .total-row td {",
            "    color: black !important;",
            "    font-weight: 700 !important;",
            "    font-size: 15px !important;",
            "    padding: 18px 20px !important;",
            "    border-top: 2px solid var(--border-color);",
            "  }",
            "  ",
            "  .subtotal-row {",
            "    background: var(--bg-secondary) !important;", # Ghost White subtotal row
            "  }",
            "  ",
            "  .subtotal-row td {",
            "    font-weight: 600 !important;",
            "    padding: 14px 20px !important;",
            "    color: var(--text-primary);",
            "  }",
            "  ",
            "  /* Footer Section */",
            "  .report-footer {",
            "    background: var(--bg-primary);",
            "    padding: 32px 24px;",
            "    text-align: center;",
            "    border-top: 1px solid var(--border-color);",
            "  }",
            "  ",
            "  .footer-content {",
            "    max-width: 600px;",
            "    margin: 0 auto;",
            "  }",
            "  ",
            "  .footer-title {",
            "    font-family: 'Manrope', sans-serif;", # Manrope for titles
            "    font-size: 18px;",
            "    font-weight: 600;",
            "    color: var(--text-primary);",
            "    margin-bottom: 10px;",
            "  }",
            "  ",
            "  .footer-text {",
            "    font-size: 12px;",
            "    color: var(--text-secondary);",
            "    line-height: 1.7;",
            "    margin-bottom: 20px;",
            "  }",
            "  ",
            "  .action-buttons {",
            "    display: flex;",
            "    gap: 12px;",
            "    justify-content: center;",
            "    flex-wrap: wrap;",
            "  }",
            "  ",
            "  .btn {",
            "    padding: 10px 24px;",
            "    border-radius: 8px;",
            "    font-weight: 600;",
            "    font-size: 13px;",
            "    cursor: pointer;",
            "    transition: all 0.3s ease;",
            "    border: 2px solid var(--primary-black);", # Outlined black button
            "    display: inline-flex;",
            "    align-items: center;",
            "    gap: 8px;",
            "    text-decoration: none;",
            "    color: var(--primary-black);",
            "    background: transparent;",
            "  }",
            "  ",
            "  .btn:hover {",
            "    background: var(--primary-black);", # Hover to solid black
            "    color: white;",
            "    transform: translateY(-1px);",
            "    box-shadow: 0 4px 8px rgba(0,0,0,0.1);",
            "  }",
            "  ",
            "  .btn-primary {",
            "    background: var(--primary-black);", # Primary is solid black
            "    color: white;",
            "    box-shadow: var(--shadow-sm);",
            "    border-color: var(--primary-black);",
            "  }",
            "  .btn-primary:hover {",
            "    background: var(--primary-light-black);", # Darker black on hover
            "    border-color: var(--primary-light-black);",
            "  }",
            "  ",
            "  /* Secondary button is the default outlined style */",
            "  .btn-secondary {",
            "    background: transparent;",
            "    color: var(--primary-black);",
            "    border-color: var(--primary-black);",
            "  }",
            "  .btn-secondary:hover {",
            "    background: var(--primary-black);",
            "    color: white;",
            "  }",
            "  ",
            "  /* Timestamp Badge */",
            "  .timestamp-badge {",
            "    display: inline-flex;",
            "    align-items: center;",
            "    gap: 6px;",
            "    background: var(--bg-accent);",
            "    padding: 6px 12px;",
            "    border-radius: 6px;",
            "    font-size: 11px;",
            "    color: var(--text-secondary);",
            "    margin-top: 20px;",
            "  }",
            "  .timestamp-badge i {",
            "    font-size: 12px;",
            "  }",
            "  ",
            "  /* Responsive Design */",
            "  @media (max-width: 1024px) {",
            "    .report-header {",
            "      padding: 28px 20px 20px;",
            "    }",
            "    ",
            "    .firm-name {",
            "      font-size: 24px;",
            "    }",
            "    ",
            "    .table-section {",
            "      padding: 28px 20px;",
            "    }",
            "  }",
            "  ",
            "  @media (max-width: 768px) {",
            "    body {",
            "      padding: 16px;",
            "    }",
            "    ",
            "    .report-card {",
            "      border-radius: 10px;",
            "    }",
            "    ",
            "    .report-header {",
            "      padding: 24px 16px 16px;",
            "    }",
            "    ",
            "    .firm-name {",
            "      font-size: 20px;",
            "    }",
            "    ",
            "    .firm-meta {",
            "      flex-direction: column;",
            "      gap: 16px;",
            "    }",
            "    ",
            "    .stats-grid {",
            "      grid-template-columns: 1fr;",
            "    }",
            "    ",
            "    .table-section {",
            "      padding: 20px 12px;",
            "    }",
            "    ",
            "    thead th,",
            "    td {",
            "      padding: 12px 14px;",
            "      font-size: 12px;",
            "    }",
            "    ",
            "    .report-footer {",
            "      padding: 24px 16px;",
            "    }",
            "    ",
            "    .action-buttons {",
            "      flex-direction: column;",
            "    }",
            "    ",
            "    .btn {",
            "      width: 100%;",
            "      justify-content: center;",
            "    }",
            "  }",
            "  ",
            "  /* Page Break Helper Class */",
            "  .page-break {",
            "    page-break-after: always;",
            "    break-after: page;",
            "    height: 0;",
            "    margin: 0;",
            "    padding: 0;",
            "  }",
            "  ",
            "  /* Print Styles */",
            "  @media print {",
            "    @page {",
            "      size: A4;",
            "      margin: 15mm 10mm 15mm 10mm;",
            "    }",
            "    ",
            "    body {",
            "      background: white;",
            "      padding: 0;",
            "      margin: 0;",
            "      font-size: 10px;",
            "    }",
            "    ",
            "    .container {",
            "      max-width: 100%;",
            "      padding: 0;",
            "    }",
            "    ",
            "    .report-card {",
            "      box-shadow: none;",
            "      border-radius: 0;",
            "    }",
            "    ",
            "    .report-header {",
            "      page-break-after: always;",
            "      min-height: 200mm;",
            "      display: flex;",
            "      align-items: center;",
            "      justify-content: center;",
            "    }",
            "    ",
            "    .report-header::before {",
            "      display: none;",
            "    }",
            "    ",
            "    .action-buttons {",
            "      display: none;",
            "    }",
            "    ",
            "    tbody tr:hover {",
            "      background: transparent;",
            "    }",
            "    ",
            "    /* Page break controls */",
            "    .section-header {",
            "      page-break-before: auto;",
            "      page-break-after: avoid;",
            "      page-break-inside: avoid;",
            "    }",
            "    ",
            "    .total-row {",
            "      page-break-before: avoid;",
            "      page-break-after: auto;",
            "      page-break-inside: avoid;",
            "    }",
            "    ",
            "    tr {",
            "      page-break-inside: avoid;",
            "    }",
            "    ",
            "    table {",
            "      page-break-inside: auto;",
            "      font-size: 9px;",
            "    }",
            "    ",
            "    td, th {",
            "      padding: 6px 8px !important;",
            "    }",
            "    ",
            "    thead {",
            "      display: table-header-group;",
            "    }",
            "    ",
            "    tfoot {",
            "      display: table-footer-group;",
            "    }",
            "    ",
            "    .page-break {",
            "      page-break-after: always;",
            "      break-after: page;",
            "    }",
            "  }",
            "  ",
            "  /* Loading Animation */",
            "  @keyframes shimmer {",
            "    0% { background-position: -1000px 0; }",
            "    100% { background-position: 1000px 0; }",
            "  }",
            "  ",
            "  .loading {",
            "    animation: shimmer 2s infinite;",
            "    background: linear-gradient(to right, #f6f7f8 0%, #edeef1 20%, #f6f7f8 40%, #f6f7f8 100%);",
            "    background-size: 1000px 100%;",
            "  }",
            "</style>",
            "</head>",
            "<body>",
            "<div class='container'>",
            "<div class='report-card'>",
            "",
            "<!-- Header Section -->",
            "<div class='report-header'>",
            "<div class='header-content'>",
            f"<span class='report-badge'><i class='fas fa-chart-line'></i> {sheet_name}</span>", # Professional Icon
            f"<h1 class='firm-name'>{firm_name or 'Financial Report'}</h1>",
            "<div class='firm-meta'>",
        ]
        
        if proprietor:
            html_parts.extend([
                "<div class='meta-item'>",
                "<span class='meta-label'>Proprietor</span>",
                f"<span class='meta-value'>{proprietor}</span>",
                "</div>",
            ])
        
        if sector:
            html_parts.extend([
                "<div class='meta-item'>",
                "<span class='meta-label'>Sector</span>",
                f"<span class='meta-value'>{sector}</span>",
                "</div>",
            ])
        
        if nature_of_business:
            html_parts.extend([
                "<div class='meta-item'>",
                "<span class='meta-label'>Nature of Business</span>",
                f"<span class='meta-value'>{nature_of_business}</span>",
                "</div>",
            ])
        
        html_parts.extend([
            "<div class='meta-item'>",
            "<span class='meta-label'>Generated</span>",
            f"<span class='meta-value'>{datetime.datetime.now().strftime('%b %d, %Y')}</span>",
            "</div>",
            "</div>",
            "</div>",
            "</div>",
            "",
            "<!-- Stats Grid -->",
        ])
        
        html_parts.extend([
            "</div>",
            "",
            "<!-- Table Section -->",
            "<div class='table-section'>",
            "<h2 class='section-title'>Financial Details</h2>",
            "<p class='section-subtitle'>Comprehensive breakdown of financial data and calculations</p>",
            "<div class='table-wrapper'>",
            "<table>",
            "<thead>",
            "<tr>",
            "<th>Particulars</th>",
            "<th>Amount</th>",
            "</tr>",
            "</thead>",
            "<tbody>",
        ])
        
        # Track row count for page breaks
        visible_row_count = 0
        PAGE_BREAK_INTERVAL = 50  # Insert page break every 50 rows
        
        # Process each row
        for row_idx in range(1, max_row + 1):
            # Skip hidden rows
            try:
                if sheet.Rows(row_idx).Hidden:
                    continue
            except:
                pass

            row_data = []
            is_header = False
            is_total = False
            is_empty_row = True
            
            # First pass: collect row data
            for col_idx in range(1, max_col + 1):
                cell = sheet.Cells(row_idx, col_idx)
                cell_value = cell.Value
                
                if cell_value is None:
                    cell_value = ""
                elif isinstance(cell_value, (int, float)):
                    if isinstance(cell_value, float):
                        if cell_value % 1 == 0:
                            cell_value = int(cell_value)
                    # Store in JSON
                    json_data["data"][f"R{row_idx}C{col_idx}"] = cell_value
                else:
                    cell_value = str(cell_value)
                    json_data["data"][f"R{row_idx}C{col_idx}"] = cell_value
                
                if cell_value != "":
                    is_empty_row = False
                
                row_data.append({
                    "value": cell_value,
                    "cell": cell,
                    "col_idx": col_idx
                })
            
            # Skip completely empty rows
            if is_empty_row:
                continue
            
            # Detect row type
            first_value = str(row_data[0]["value"]).lower() if row_data else ""
            if any(keyword in first_value for keyword in ["step", "financials", "ratios", "particulars", "profit", "balance", "sheet", "statement"]):
                is_header = True
            elif any(keyword in first_value for keyword in ["total", "net", "grand"]):
                is_total = True
            # Cells will be editable when unlocked in Excel; no need for inline markers
            # Build row HTML
            row_class = ""
            if is_header:
                row_class = " class='section-header'"
            elif is_total:
                row_class = " class='total-row'"
            elif "subtotal" in first_value or "sub-total" in first_value:
                row_class = " class='subtotal-row'"
            
            html_parts.append(f"  <tr{row_class}>")
            
            for cell_data in row_data:
                cell = cell_data["cell"]
                cell_value = cell_data["value"]
                col_idx = cell_data["col_idx"]
                
                # Determine cell class
                cell_classes = []
                if col_idx == 1:
                    cell_classes.append("item-name")
                else:
                    cell_classes.append("item-value")
                
                # Format numeric values as currency
                formatted_value = cell_value
                if isinstance(cell_value, (int, float)) and cell_value != "" and col_idx > 1:
                    cell_classes.append("currency")
                    # Format with commas but without currency symbol (CSS will add it)
                    if isinstance(cell_value, float):
                        formatted_value = f"{cell_value:,.2f}"
                    else:
                        formatted_value = f"{cell_value:,}"
                
                # Basic styling from cell
                style_parts = []
                
                # Background color - Only apply if NOT a special row (let CSS handle it)
                if not is_header and not is_total and "subtotal" not in first_value and "sub-total" not in first_value:
                    try:
                        interior_color = cell.Interior.Color
                        if interior_color != 16777215:  # Not white
                            r = interior_color & 255
                            g = (interior_color >> 8) & 255
                            b = (interior_color >> 16) & 255
                            style_parts.append(f"background-color: rgb({r},{g},{b})")
                    except:
                        pass
                
                # Font color (only if custom style not applied)
                try:
                    if not is_header and not is_total:
                        font_color = cell.Font.Color
                        if font_color != 0:  # Not black
                            r = font_color & 255
                            g = (font_color >> 8) & 255
                            b = (font_color >> 16) & 255
                            style_parts.append(f"color: rgb({r},{g},{b})")
                except:
                    pass
                
                # Font weight
                try:
                    if cell.Font.Bold and not is_header and not is_total:
                        style_parts.append("font-weight: bold")
                except:
                    pass
                
                style_attr = "; ".join(style_parts) if style_parts else ""
                class_attr = " ".join(cell_classes) if cell_classes else ""
                
                # Handle merged cells
                merge_attrs = ""
                try:
                    merge_area = cell.MergeArea
                    if merge_area.Cells.Count > 1:
                        rowspan = merge_area.Rows.Count
                        colspan = merge_area.Columns.Count
                        if cell.Row == merge_area.Row and cell.Column == merge_area.Column:
                            if rowspan > 1:
                                merge_attrs += f" rowspan='{rowspan}'"
                            if colspan > 1:
                                merge_attrs += f" colspan='{colspan}'"
                        else:
                            continue
                except:
                    pass
                
                # Output cell
                attrs = []
                if class_attr:
                    attrs.append(f"class='{class_attr}'")
                if style_attr:
                    attrs.append(f"style='{style_attr}'")
                attrs.append(f"data-cell=\"R{row_idx}C{col_idx}\"")
                attrs.append(f"data-sheet=\"{actual_sheet_name}\"")
                is_cell_unlocked = False
                try:
                    # Excel sets Locked=True by default; unlocked cells are allowed to edit
                    is_cell_unlocked = not bool(cell.Locked)
                except Exception:
                    is_cell_unlocked = False

                if is_cell_unlocked and col_idx > 1:
                    attrs.append("data-editable=\"true\"")
                
                attr_str = " " + " ".join(attrs) if attrs else ""
                html_parts.append(f"    <td{attr_str}{merge_attrs}>{formatted_value}</td>")
            
            html_parts.append("  </tr>")
            
            # Increment visible row counter
            visible_row_count += 1
            
            # Insert page break after every PAGE_BREAK_INTERVAL rows (but not on section headers)
            if visible_row_count > 0 and visible_row_count % PAGE_BREAK_INTERVAL == 0 and not is_header:
                html_parts.append("</tbody></table></div>")
                html_parts.append("<div class='page-break'></div>")
                html_parts.append("<div class='table-wrapper'><table>")
                html_parts.append("<thead><tr><th>Particulars</th><th>Amount</th></tr></thead>")
                html_parts.append("<tbody>")
        
        html_parts.extend([
            "</tbody>",
            "</table>",
            "</div>",
            "</div>",
            "",
            "<!-- Footer Section -->",
            "<div class='report-footer'>",
            "<div class='footer-content'>",
            f"<h3 class='footer-title'><i class='fas fa-check-circle'></i> Report Generated Successfully</h3>", # Professional Icon
            "</div>",
            "<div class='timestamp-badge'>",
            f"<i class='fas fa-clock'></i> Generated on " + datetime.datetime.now().strftime('%B %d, %Y at %I:%M %p'), # Professional Icon
            "</div>",
            "</div>",
            "</div>",
            "",
            "</div>",
            "</div>",
            "",
            "<script>",
            "// Store JSON data for programmatic access",
            f"window.reportData = {json.dumps(json_data, ensure_ascii=False)};",
            "",
            "console.log('%c📊 Financial Report Data Loaded', 'color: #7c3aed; font-weight: bold; font-size: 16px; font-family: Inter, sans-serif;');",
            "console.log('%c━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━', 'color: #7c3aed;');",
            "console.log('%c📄 Sheet Name:', 'color: #6b7280; font-weight: 600;', window.reportData.sheetName);",
            "console.log('%c🔢 Total Cells:', 'color: #6b7280; font-weight: 600;', Object.keys(window.reportData.data).length);",
            "console.log('%c⏰ Timestamp:', 'color: #6b7280; font-weight: 600;', window.reportData.timestamp);",
            "console.log('%c━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━', 'color: #7c3aed;');",
            "console.log('%c💡 Access data: window.reportData.data[\"R1C1\"]', 'color: #10b981; font-style: italic;');",
            "",
            "// Download report as PDF (placeholder function)",
            "function downloadReport() {",
            "  alert('PDF download functionality will be implemented by the backend. (Ctrl+P to print)');", # Updated alert
            "  console.log('Download request initiated for:', window.reportData.sheetName);",
            "}",
            "",
            "// Add smooth scroll behavior",
            "document.querySelectorAll('a[href^=\"#\"]').forEach(anchor => {",
            "  anchor.addEventListener('click', function (e) {",
            "    e.preventDefault();",
            "    const target = document.querySelector(this.getAttribute('href'));",
            "    if (target) {",
            "      target.scrollIntoView({ behavior: 'smooth', block: 'start' });",
            "    }",
            "  });",
            "});",
            "",
            "// Add loading state handler",
            "window.addEventListener('load', () => {",
            "  document.querySelectorAll('.loading').forEach(el => {",
            "    el.classList.remove('loading');",
            "  });",
            "});",
            "",
            "// Add table row highlight on click",
            "document.querySelectorAll('tbody tr').forEach(row => {",
            "  row.addEventListener('click', function() {",
            "    document.querySelectorAll('tbody tr').forEach(r => {",
            "      r.style.outline = 'none';",
            "    });",
            "    this.style.outline = '2px solid var(--border-color)';", # Highlight with neutral border color
            "    this.style.outlineOffset = '-2px';",
            "  });",
            "});",
            "",
            "// Add keyboard navigation",
            "document.addEventListener('keydown', (e) => {",
            "  if (e.ctrlKey && e.key === 'p') {",
            "    e.preventDefault();",
            "    window.print();",
            "  }",
            "});",
            "",
            "// Performance monitoring",
            "if (window.performance) {",
            "  const perfData = window.performance.timing;",
            "  const pageLoadTime = perfData.loadEventEnd - perfData.navigationStart;",
            "  console.log('%c⚡ Page Load Time:', 'color: #10b981; font-weight: 600;', pageLoadTime + 'ms');",
            "}",
            "",
            "// Add animation observer for elements",
            "const observerOptions = {",
            "  threshold: 0.1,",
            "  rootMargin: '0px 0px -50px 0px'",
            "};",
            "",
            "const observer = new IntersectionObserver((entries) => {",
            "  entries.forEach(entry => {",
            "    if (entry.isIntersecting) {",
            "      entry.target.style.opacity = '1';",
            "      entry.target.style.transform = 'translateY(0)';",
            "    }",
            "  });",
            "}, observerOptions);",
            "",
            "document.querySelectorAll('.stat-card, .table-wrapper').forEach(el => {",
            "  el.style.opacity = '0';",
            "  el.style.transform = 'translateY(20px)';",
            "  el.style.transition = 'opacity 0.6s ease, transform 0.6s ease';",
            "  observer.observe(el);",
            "});",
            "</script>",
            "</body>",
            "</html>"
        ])
        
        html_content = "\n".join(html_parts)
        print(f"[HTML COM Generator] SUCCESS: HTML generated successfully ({len(html_content)} chars)", file=sys.stderr)
        print(f"[HTML COM Generator] SUCCESS: JSON data extracted ({len(json_data['data'])} cells)", file=sys.stderr)
        
        # Clean up
        wb.Close(False)
        excel.Quit()
        
        return html_content, json_data
        
    except Exception as e:
        print(f"[HTML COM Generator] ❌ Error: {str(e)}", file=sys.stderr)
        import traceback
        traceback.print_exc(file=sys.stderr)
        try:
            if 'wb' in locals():
                wb.Close(False)
            if 'excel' in locals():
                excel.Quit()
        except:
            pass
        return ""

def generate_html_from_excel_sheet(excel_path: str, sheet_name: str, header_data: dict = None):
    """
    Convert an Excel sheet to HTML with complete styling preservation.
    Returns tuple: (html_content, json_data) for both COM and fallback methods.
    FALLBACK NOW PROPERLY EVALUATES FORMULAS AND DISPLAYS VALUES.
    """
    try:
        print(f"[HTML Generator] Loading workbook: {excel_path}", file=sys.stderr)
        print(f"[HTML Generator] COM_AVAILABLE: {COM_AVAILABLE}", file=sys.stderr)
        
        # Try using win32com first to get calculated values
        html_content = None
        json_data = {}
        if COM_AVAILABLE:
            try:
                print(f"[HTML Generator] Attempting to use Excel COM method", file=sys.stderr)
                html_content, json_data = generate_html_from_excel_com(excel_path, sheet_name, header_data=header_data)
                if html_content:
                    print(f"[HTML Generator] ✓ Successfully generated HTML using COM method", file=sys.stderr)
                    return html_content, json_data
                else:
                    print(f"[HTML Generator] ⚠ COM method returned empty content - may indicate sheet not found or COM error", file=sys.stderr)
            except Exception as com_error:
                print(f"[HTML Generator] ⚠ COM method failed, falling back to openpyxl: {com_error}", file=sys.stderr)
                import traceback
                traceback.print_exc(file=sys.stderr)
        else:
            print(f"[HTML Generator] COM not available, using openpyxl fallback", file=sys.stderr)
        
        # ========================================
        # FALLBACK METHOD WITH PROPER VALUE HANDLING
        # ========================================
        print(f"[HTML Generator] Using openpyxl fallback method WITH PROFESSIONAL STYLING", file=sys.stderr)
        
        from decimal import Decimal
        import math
        
        print(f"[HTML Generator] Reading workbook with openpyxl (data_only=True) to capture evaluated values...", file=sys.stderr)
        values_wb = None
        structure_wb = None
        
        try:
            values_wb = load_workbook(excel_path, data_only=True)
            structure_wb = load_workbook(excel_path, data_only=False)
            available_sheets = values_wb.sheetnames
            actual_sheet_name = find_sheet_match(sheet_name, available_sheets)
            if not actual_sheet_name:
                print(f"[HTML Generator] ERROR: Sheet '{sheet_name}' not found", file=sys.stderr)
                print(f"[HTML Generator] Available sheets: {available_sheets}", file=sys.stderr)
                return "", {}
            
            values_sheet = values_wb[actual_sheet_name]
            sheet = structure_wb[actual_sheet_name]
            max_row = values_sheet.max_row or 0
            max_col = values_sheet.max_column or 0
            print(
                f"[HTML Generator] Processing sheet: {actual_sheet_name} (matched from '{sheet_name}')",
                file=sys.stderr
            )
            print(
                f"[HTML Generator] Worksheet size detected: {max_row} rows x {max_col} columns",
                file=sys.stderr
            )
            
            json_data = {
                "sheetName": actual_sheet_name,
                "data": {},
                "timestamp": datetime.datetime.now().isoformat()
            }

            merged_ranges = {}
            for merged_range in sheet.merged_cells.ranges:
                merged_ranges[(merged_range.min_row, merged_range.min_col)] = {
                    'rowspan': merged_range.max_row - merged_range.min_row + 1,
                    'colspan': merged_range.max_col - merged_range.min_col + 1
                }
            
            def _to_native_excel_value(raw_value):
                """Convert xlcalculator ExcelType or other wrappers to native Python."""
                if raw_value is None:
                    return None
                try:
                    if hasattr(raw_value, 'to_python'):
                        return raw_value.to_python()
                    if hasattr(raw_value, 'value') and not isinstance(raw_value, (str, bytes)):
                        # ExcelType exposes .value containing the typed payload
                        nested_val = raw_value.value
                        if hasattr(nested_val, 'to_python'):
                            return nested_val.to_python()
                        return nested_val
                except Exception as conv_err:
                    print(f"[HTML Generator] Warning: Could not convert ExcelType value ({conv_err})", file=sys.stderr)
                return raw_value

            def normalize_cell_value(raw_value):
                if raw_value is None:
                    return ""
                raw_value = _to_native_excel_value(raw_value)
                if isinstance(raw_value, str):
                    return raw_value.strip()
                if isinstance(raw_value, Decimal):
                    return float(raw_value)
                if isinstance(raw_value, float):
                    if math.isnan(raw_value) or math.isinf(raw_value):
                        return ""
                    return float(raw_value)
                return raw_value
            
            def serialize_for_json(value):
                if isinstance(value, (datetime.datetime, datetime.date)):
                    return value.isoformat()
                return value
            
            def extract_text_value(row_num: int, col_num: int) -> str:
                try:
                    cell_val = values_sheet.cell(row=row_num, column=col_num).value
                except Exception:
                    cell_val = None
                if cell_val is None:
                    return ""
                return str(cell_val).strip()
            
            firm_name = extract_text_value(3, 2)
            proprietor = ""
            sector = ""
            nature_of_business = ""
            
            # Use provided header data if available
            if header_data:
                proprietor = header_data.get('proprietor', '')
                sector = header_data.get('sector', '')
                nature_of_business = header_data.get('natureOfBusiness', '')
                
            # Fallback to Excel extraction if not provided
            if not proprietor:
                proprietor = extract_text_value(4, 2)
            if not sector:
                sector = extract_text_value(6, 2)
            if not nature_of_business:
                nature_of_business = extract_text_value(7, 2)
            
            calc_evaluator = None
            if XL_CALC_AVAILABLE:
                try:
                    print("[HTML Generator] Initializing xlcalculator model for formula evaluation", file=sys.stderr)
                    compiler = ModelCompiler()
                    parsed_model = compiler.read_and_parse_archive(excel_path)
                    _ensure_custom_xl_functions_registered()
                    _coerce_model_constants_to_excel_types(parsed_model)
                    calc_evaluator = Evaluator(parsed_model)
                    print("[HTML Generator] xlcalculator model ready", file=sys.stderr)
                except Exception as calc_init_error:
                    calc_evaluator = None
                    print(f"[HTML Generator] Warning: xlcalculator unavailable ({calc_init_error})", file=sys.stderr)
            else:
                print("[HTML Generator] xlcalculator package not available; complex formulas may remain blank", file=sys.stderr)

            sheet_matrix: List[List[Any]] = []
            non_empty_cells = 0
            calc_debug_count = 0
            sheet_prefix = (
                f"'{actual_sheet_name}'" if any(ch in actual_sheet_name for ch in (" ", "-", "."))
                else actual_sheet_name
            )

            for row_idx in range(1, max_row + 1):
                # Skip hidden rows
                try:
                    if sheet.row_dimensions[row_idx].hidden:
                        continue
                except:
                    pass

                row_values: List[Any] = []
                for col_idx in range(1, max_col + 1):
                    raw_value = values_sheet.cell(row=row_idx, column=col_idx).value
                    normalized_value = normalize_cell_value(raw_value)
                    structure_cell = sheet.cell(row=row_idx, column=col_idx)
                    cell_formula = structure_cell.value
                    has_formula = False
                    try:
                        has_formula = (
                            structure_cell.data_type == "f"
                            or (isinstance(cell_formula, str) and cell_formula.startswith("="))
                        )
                    except Exception:
                        has_formula = False

                    if calc_evaluator and (normalized_value in ("", None) or has_formula):
                        cell_ref = f"{get_column_letter(col_idx)}{row_idx}"
                        sheet_ref = f"{sheet_prefix}!{cell_ref}"
                        try:
                            calc_value = calc_evaluator.evaluate(sheet_ref)
                            normalized_calc_value = normalize_cell_value(calc_value)
                            if normalized_calc_value not in ("", None):
                                normalized_value = normalized_calc_value
                                if calc_debug_count < 6:
                                    source = "formula" if has_formula else "blank"
                                    print(
                                        f"[HTML Generator] xlcalculator filled {sheet_ref} ({source}) => {normalized_value}",
                                        file=sys.stderr
                                    )
                                    calc_debug_count += 1
                        except KeyError:
                            # Missing sheet/cell references can be safely ignored
                            pass
                        except Exception as calc_eval_error:
                            if calc_debug_count < 12:
                                print(
                                    f"[HTML Generator] xlcalculator skip {sheet_ref}: {calc_eval_error}",
                                    file=sys.stderr
                                )
                                calc_debug_count += 1

                    row_values.append(normalized_value)

                    if normalized_value not in ("", None):
                        json_key = f"R{row_idx}C{col_idx}"
                        json_data["data"][json_key] = serialize_for_json(normalized_value)
                        non_empty_cells += 1

                sheet_matrix.append((row_idx, row_values))
            
            total_rows = len(sheet_matrix)
            total_cols = max_col
            print(
                f"[HTML Generator] Prepared matrix with {total_rows} rows for HTML conversion",
                file=sys.stderr
            )
            print(
                f"[HTML Generator] Non-empty cells captured for JSON: {non_empty_cells}",
                file=sys.stderr
            )
            sample_preview = list(json_data["data"].items())[:10]
            if sample_preview:
                print("[HTML Generator] Sample JSON pairs:", file=sys.stderr)
                for key, value in sample_preview:
                    print(f"  - {key} => {value}", file=sys.stderr)
            else:
                print("[HTML Generator] JSON data is currently empty", file=sys.stderr)
        finally:
            if values_wb:
                values_wb.close()
            if structure_wb:
                structure_wb.close()
        
        # Build HTML with EXACT SAME professional styling as COM method
        html_parts = [
            "<!DOCTYPE html>",
            "<html lang='en'>",
            "<head>",
            "<meta charset='UTF-8'>",
            "<meta name='viewport' content='width=device-width, initial-scale=1.0'>",
            f"<title>Financial Report - {firm_name or sheet_name}</title>",
            "<link href='https://fonts.googleapis.com/css2?family=Manrope:wght@400;500;600;700;800&family=Inter:wght@300;400;500;600;700&display=swap' rel='stylesheet'>",
            "<link rel='stylesheet' href='https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.0.0-beta3/css/all.min.css'>",
            "<style>",
            "  :root {",
            "    --primary-purple: #7c3aed;",
            "    --primary-dark-purple: #6d28d9;",
            "    --primary-black: #1f2937;",
            "    --primary-light-black: #374151;",
            "    --ghost-white: #F8F8FF;",
            "    --success-green: #10b981;",
            "    --text-primary: var(--primary-black);",
            "    --text-secondary: #6b7280;",
            "    --bg-primary: #ffffff;",
            "    --bg-secondary: var(--ghost-white);",
            "    --bg-accent: #e5e7eb;",
            "    --border-color: #e5e7eb;",
            "    --shadow-soft: 0 2px 15px -3px rgba(0, 0, 0, 0.07), 0 10px 20px -2px rgba(0, 0, 0, 0.04);",
            "    --shadow-md: 0 4px 6px -1px rgba(0, 0, 0, 0.1);",
            "    --shadow-lg: 0 10px 15px -3px rgba(0, 0, 0, 0.1);",
            "    --shadow-xl: 0 20px 25px -5px rgba(0, 0, 0, 0.1);",
            "  }",
            "  ",
            "  * {",
            "    margin: 0;",
            "    padding: 0;",
            "    box-sizing: border-box;",
            "  }",
            "  ",
            "  body {",
            "    font-family: 'Inter', -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;",
            "    background-color: var(--bg-secondary);",
            "    min-height: 100vh;",
            "    padding: 20px;",
            "    line-height: 1.6;",
            "    color: var(--text-primary);",
            "    -webkit-font-smoothing: antialiased;",
            "    -moz-osx-font-smoothing: grayscale;",
            "  }",
            "  ",
            "  .container {",
            "    max-width: 1400px;",
            "    margin: 0 auto;",
            "  }",
            "  ",
            "  .report-card {",
            "    background: var(--bg-primary);",
            "    border-radius: 12px;",
            "    box-shadow: var(--shadow-soft);",
            "    overflow: hidden;",
            "    animation: slideUp 0.6s ease-out;",
            "  }",
            "  ",
            "  @keyframes slideUp {",
            "    from {",
            "      opacity: 0;",
            "      transform: translateY(30px);",
            "    }",
            "    to {",
            "      opacity: 1;",
            "      transform: translateY(0);",
            "    }",
            "  }",
            "  ",
            "  /* Header Section */",
            "  .report-header {",
            "    padding: 32px 24px;",
            "    color: black;",
            "    position: relative;",
            "    overflow: hidden;",
            "    border-bottom: 1px solid rgba(255, 255, 255, 0.1);",
            "  }",
            "  ",
            "  .report-header::before {",
            "    content: '';",
            "    position: absolute;",
            "    top: 0;",
            "    right: 0;",
            "    width: 200px;",
            "    height: 200px;",
            "    background: radial-gradient(circle, rgba(255,255,255,0.15) 0%, transparent 70%);",
            "    border-radius: 50%;",
            "    transform: translate(30%, -30%);",
            "  }",
            "  ",
            "  .header-content {",
            "    position: relative;",
            "    z-index: 1;",
            "  }",
            "  ",
            "  .report-badge {",
            "    display: inline-flex;",
            "    align-items: center;",
            "    gap: 8px;",
            "    background: rgba(255, 255, 255, 0.2);",
            "    backdrop-filter: blur(5px);",
            "    padding: 6px 16px;",
            "    border-radius: 50px;",
            "    font-size: 12px;",
            "    font-weight: 600;",
            "    letter-spacing: 0.5px;",
            "    text-transform: uppercase;",
            "    margin-bottom: 24px;",
            "  }",
            "  ",
            "  .firm-name {",
            "    font-size: 28px;",
            "    font-weight: 700;",
            "    margin-bottom: 16px;",
            "    letter-spacing: -0.5px;",
            "    color: var(--primary-black);",
            "  }",
            "  ",
            "  .firm-meta {",
            "    display: flex;",
            "    flex-wrap: wrap;",
            "    gap: 24px;",
            "    margin-top: 16px;",
            "    padding-top: 0;",
            "    border-top: none;",
            "  }",
            "  ",
            "  .meta-item {",
            "    display: flex;",
            "    flex-direction: column;",
            "    gap: 4px;",
            "  }",
            "  ",
            "  .meta-label {",
            "    font-size: 11px;",
            "    font-weight: 600;",
            "    color: var(--text-secondary);",
            "    text-transform: uppercase;",
            "    letter-spacing: 1px;",
            "  }",
            "  ",
            "  .meta-value {",
            "    font-size: 15px;",
            "    font-weight: 600;",
            "    color: var(--primary-black);",
            "  }",
            "  ",
            "  /* Stats Grid Section */",
            "  .stats-grid {",
            "    display: grid;",
            "    grid-template-columns: repeat(auto-fit, minmax(150px, 1fr));",
            "    gap: 1px;",
            "    background: var(--border-color);",
            "    border-top: 1px solid var(--border-color);",
            "  }",
            "  ",
            "  .stat-card {",
            "    background: var(--bg-primary);",
            "    padding: 24px;",
            "    text-align: center;",
            "    transition: all 0.2s ease;",
            "  }",
            "  ",
            "  .stat-card:hover {",
            "    background: var(--bg-secondary);",
            "  }",
            "  ",
            "  .stat-icon {",
            "    font-size: 20px;",
            "    margin-bottom: 10px;",
            "  }",
            "  ",
            "  .stat-label {",
            "    font-size: 11px;",
            "    font-weight: 600;",
            "    color: var(--text-secondary);",
            "    text-transform: uppercase;",
            "    letter-spacing: 1px;",
            "    margin-bottom: 8px;",
            "  }",
            "  ",
            "  .stat-value {",
            "    font-size: 16px;",
            "    font-weight: 700;",
            "    color: var(--primary-black);",
            "  }",
            "  ",
            "  /* Table Section */",
            "  .table-section {",
            "    padding: 32px 24px;",
            "  }",
            "  ",
            "  .section-title {",
            "    font-size: 20px;",
            "    font-weight: 700;",
            "    color: var(--primary-black);",
            "    margin-bottom: 8px;",
            "  }",
            "  ",
            "  .section-subtitle {",
            "    font-size: 13px;",
            "    color: var(--text-secondary);",
            "    margin-bottom: 24px;",
            "  }",
            "  ",
            "  .table-wrapper {",
            "    overflow-x: auto;",
            "    border: 1px solid var(--border-color);",
            "    border-radius: 8px;",
            "  }",
            "  ",
            "  table {",
            "    width: 100%;",
            "    border-collapse: collapse;",
            "    background: var(--bg-primary);",
            "  }",
            "  ",
            "  thead {",
            "    background: var(--bg-accent);",
            "    position: sticky;",
            "    top: 0;",
            "    z-index: 10;",
            "  }",
            "  ",
            "  thead th {",
            "    padding: 16px 12px;",
            "    text-align: left;",
            "    font-size: 11px;",
            "    font-weight: 700;",
            "    color: var(--primary-black);",
            "    text-transform: uppercase;",
            "    letter-spacing: 1px;",
            "    border-bottom: 1px solid var(--border-color);",
            "  }",
            "  ",
            "  thead th:last-child {",
            "    text-align: right;",
            "  }",
            "  ",
            "  tbody tr {",
            "    border-bottom: 1px solid var(--border-color);",
            "  }",
            "  ",
            "  tbody tr:hover {",
            "    background: var(--bg-secondary);",
            "  }",
            "  ",
            "  tbody tr:last-child {",
            "    border-bottom: none;",
            "  }",
            "  ",
            "  td {",
            "    padding: 14px 12px;",
            "    font-size: 13px;",
            "    color: var(--text-primary);",
            "  }",
            "  ",
            "  .item-name {",
            "    font-weight: 500;",
            "    max-width: 400px;",
            "  }",
            "  ",
            "  .item-value {",
            "    text-align: right;",
            "    font-weight: 600;",
            "    font-family: 'SF Mono', 'Monaco', 'Courier New', monospace;",
            "    font-size: 13px;",
            "  }",
            "  ",
            "  .currency::before {",
            "    content: '₹ ';",
            "    color: var(--text-secondary);",
            "    margin-right: 2px;",
            "  }",
            "  ",
            "  /* Row Types */",
            "  .section-header {",
            "    background: var(--bg-accent) !important;",
            "  }",
            "  ",
            "  .section-header td {",
            "    font-weight: 700 !important;",
            "    font-size: 12px !important;",
            "    color: var(--primary-black) !important;",
            "    padding: 14px 12px !important;",
            "    text-transform: uppercase;",
            "    letter-spacing: 1px;",
            "  }",
            "  ",
            "  .total-row {",
            "    background: var(--bg-accent) !important;",
            "  }",
            "  ",
            "  .total-row td {",
            "    color: var(--primary-black) !important;",
            "    font-weight: 700 !important;",
            "    font-size: 14px !important;",
            "    padding: 18px 12px !important;",
            "    border-top: 1px solid var(--border-color);",
            "  }",
            "  ",
            "  .subtotal-row {",
            "    background: var(--bg-accent) !important;",
            "  }",
            "  ",
            "  .subtotal-row td {",
            "    font-weight: 600 !important;",
            "    padding: 14px 12px !important;",
            "    color: var(--primary-black);",
            "  }",
            "  ",
            "  /* Footer Section */",
            "  .report-footer {",
            "    background: var(--bg-accent);",
            "    padding: 32px 24px;",
            "    text-align: center;",
            "    border-top: 1px solid var(--border-color);",
            "  }",
            "  ",
            "  .footer-content {",
            "    max-width: 600px;",
            "    margin: 0 auto;",
            "  }",
            "  ",
            "  .footer-title {",
            "    font-size: 16px;",
            "    font-weight: 700;",
            "    color: var(--primary-black);",
            "    margin-bottom: 12px;",
            "  }",
            "  ",
            "  .footer-text {",
            "    font-size: 13px;",
            "    color: var(--text-secondary);",
            "    line-height: 1.6;",
            "    margin-bottom: 20px;",
            "  }",
            "  ",
            "  /* Action Buttons */",
            "  .action-buttons {",
            "    display: flex;",
            "    gap: 12px;",
            "    justify-content: center;",
            "    flex-wrap: wrap;",
            "  }",
            "  ",
            "  .btn {",
            "    padding: 10px 24px;",
            "    border-radius: 8px;",
            "    font-weight: 600;",
            "    font-size: 13px;",
            "    cursor: pointer;",
            "    transition: all 0.2s ease;",
            "    border: none;",
            "    display: inline-flex;",
            "    align-items: center;",
            "    gap: 8px;",
            "    text-decoration: none;",
            "  }",
            "  ",
            "  .btn-primary {",
            "    background: var(--primary-black);",
            "    color: white;",
            "    box-shadow: 0 2px 8px rgba(0, 0, 0, 0.1);",
            "  }",
            "  ",
            "  .btn-primary:hover {",
            "    transform: translateY(-2px);",
            "    box-shadow: 0 4px 12px rgba(0, 0, 0, 0.15);",
            "  }",
            "  ",
            "  .btn-secondary {",
            "    background: var(--bg-primary);",
            "    color: var(--primary-black);",
            "    border: 1px solid var(--border-color);",
            "  }",
            "  ",
            "  .btn-secondary:hover {",
            "    background: var(--bg-secondary);",
            "  }",
            "  ",
            "  /* Timestamp Badge */",
            "  .timestamp-badge {",
            "    display: inline-flex;",
            "    align-items: center;",
            "    gap: 8px;",
            "    background: var(--bg-primary);",
            "    padding: 8px 16px;",
            "    border-radius: 6px;",
            "    font-size: 12px;",
            "    color: var(--text-secondary);",
            "    border: 1px solid var(--border-color);",
            "  }",
            "  ",
            "  /* Responsive Design */",
            "  @media (max-width: 1024px) {",
            "    .report-header {",
            "      padding: 24px 18px;",
            "    }",
            "    ",
            "    .firm-name {",
            "      font-size: 24px;",
            "    }",
            "    ",
            "    .stats-grid {",
            "      grid-template-columns: repeat(auto-fit, minmax(120px, 1fr));",
            "    }",
            "    ",
            "    .table-section {",
            "      padding: 24px 18px;",
            "    }",
            "  }",
            "  ",
            "  @media (max-width: 768px) {",
            "    body {",
            "      padding: 12px;",
            "    }",
            "    ",
            "    .report-card {",
            "      border-radius: 8px;",
            "    }",
            "    ",
            "    .report-header {",
            "      padding: 20px 16px;",
            "    }",
            "    ",
            "    .firm-name {",
            "      font-size: 20px;",
            "      margin-bottom: 12px;",
            "    }",
            "    ",
            "    .firm-meta {",
            "      gap: 16px;",
            "      margin-top: 12px;",
            "    }",
            "    ",
            "    .stats-grid {",
            "      grid-template-columns: repeat(2, 1fr);",
            "    }",
            "    ",
            "    .table-section {",
            "      padding: 18px 12px;",
            "    }",
            "    ",
            "    .section-title {",
            "      font-size: 18px;",
            "    }",
            "    ",
            "    thead th, td {",
            "      padding: 12px 8px;",
            "      font-size: 12px;",
            "    }",
            "    ",
            "    .report-footer {",
            "      padding: 24px 16px;",
            "    }",
            "    ",
            "    .action-buttons {",
            "      flex-direction: column;",
            "    }",
            "    ",
            "    .btn {",
            "      width: 100%;",
            "      justify-content: center;",
            "    }",
            "  }",
            "  ",
            "  @media (max-width: 480px) {",
            "    .stats-grid {",
            "      grid-template-columns: 1fr;",
            "    }",
            "    ",
            "    .firm-meta {",
            "      flex-direction: column;",
            "      gap: 12px;",
            "    }",
            "  }",
            "  ",
            "  @media print {",
            "    body {",
            "      background: white;",
            "      padding: 0;",
            "    }",
            "    ",
            "    .report-card {",
            "      box-shadow: none;",
            "      border-radius: 0;",
            "    }",
            "    ",
            "    .report-header::before {",
            "      display: none;",
            "    }",
            "    ",
            "    .action-buttons {",
            "      display: none;",
            "    }",
            "    ",
            "    /* Page break controls */",
            "    .section-header {",
            "      page-break-before: auto;",
            "      page-break-after: avoid;",
            "      page-break-inside: avoid;",
            "    }",
            "    ",
            "    .total-row {",
            "      page-break-before: avoid;",
            "      page-break-after: auto;",
            "      page-break-inside: avoid;",
            "    }",
            "    ",
            "    tr {",
            "      page-break-inside: avoid;",
            "    }",
            "    ",
            "    table {",
            "      page-break-inside: auto;",
            "    }",
            "    ",
            "    thead {",
            "      display: table-header-group;",
            "    }",
            "    ",
            "    tfoot {",
            "      display: table-footer-group;",
            "    }",
            "  }",
            "</style>",
            "</head>",
            "<body>",
            "<div class='container'>",
            "<div class='report-card'>",
            "",
            "<!-- Header Section -->",
            "<div class='report-header'>",
            "<div class='header-content'>",
            f"<span class='report-badge'>📊 {sheet_name}</span>",
            f"<h1 class='firm-name'>{firm_name or 'Financial Report'}</h1>",
            "<div class='firm-meta'>",
        ]
        
        if proprietor:
            html_parts.extend([
                "<div class='meta-item'>",
                "<span class='meta-label'>Proprietor</span>",
                f"<span class='meta-value'>{proprietor}</span>",
                "</div>",
            ])
        
        if sector:
            html_parts.extend([
                "<div class='meta-item'>",
                "<span class='meta-label'>Sector</span>",
                f"<span class='meta-value'>{sector}</span>",
                "</div>",
            ])
        
        if nature_of_business:
            html_parts.extend([
                "<div class='meta-item'>",
                "<span class='meta-label'>Nature of Business</span>",
                f"<span class='meta-value'>{nature_of_business}</span>",
                "</div>",
            ])
        
        html_parts.extend([
            "<div class='meta-item'>",
            "<span class='meta-label'>Generated</span>",
            f"<span class='meta-value'>{datetime.datetime.now().strftime('%b %d, %Y')}</span>",
            "</div>",
            "</div>",
            "</div>",
            "</div>",
            "",
            "<!-- Table Section -->",
            "<div class='table-section'>",
            "<h2 class='section-title'>Financial Details</h2>",
            "<p class='section-subtitle'>Comprehensive breakdown of financial data and calculations</p>",
            "<div class='table-wrapper'>",
            "<table>",
            "<thead>",
            "<tr>",
            "<th>Particulars</th>",
            "<th>Amount</th>",
            "</tr>",
            "</thead>",
            "<tbody>",
        ])
        
        print(f"[HTML Generator] Processing {total_rows} rows from evaluated worksheet data", file=sys.stderr)

        def format_cell_for_html(value, col_idx):
            if value in ("", None):
                return ""
            if isinstance(value, (datetime.datetime, datetime.date)):
                # Present dates in a reader-friendly format
                try:
                    return value.strftime('%d %b %Y')
                except Exception:
                    return value.isoformat()
            if isinstance(value, float):
                if col_idx > 1:
                    return f"{value:,.2f}"
                return f"{value:.2f}".rstrip('0').rstrip('.')
            if isinstance(value, int):
                if col_idx > 1:
                    return f"{value:,}"
                return str(value)
            return str(value)
        
        # Process each row from the prepared matrix
        for matrix_idx in range(total_rows):
            orig_row_idx, row_values_list = sheet_matrix[matrix_idx]
            
            row_data = []
            is_empty_row = True
            
            # Get all column values for this row
            for col_idx in range(total_cols):
                try:
                    cell_value = row_values_list[col_idx]
                except IndexError:
                    cell_value = ""
                
                if cell_value not in ("", None):
                    is_empty_row = False
                
                row_data.append({
                    "value": cell_value,
                    "col_idx": col_idx + 1  # 1-indexed for HTML
                })
            
            # Skip completely empty rows
            if is_empty_row:
                continue
            
            # Detect row type
            first_value = str(row_data[0]["value"]).lower() if row_data else ""
            is_header = any(kw in first_value for kw in ["step", "financials", "ratios", "particulars", "profit", "balance", "sheet", "statement"])
            is_total = any(kw in first_value for kw in ["total", "net", "grand"])
            is_subtotal = "subtotal" in first_value or "sub-total" in first_value
            
            row_class = ""
            if is_header:
                row_class = " class='section-header'"
            elif is_total:
                row_class = " class='total-row'"
            elif is_subtotal:
                row_class = " class='subtotal-row'"
            
            html_parts.append(f"  <tr{row_class}>")
            
            for cell_data in row_data:
                cell_value = cell_data["value"]
                col_idx = cell_data["col_idx"]
                
                # Cell classes
                cell_classes = []
                if col_idx == 1:
                    cell_classes.append("item-name")
                else:
                    cell_classes.append("item-value")
                
                # Format numeric values as currency when needed
                formatted_value = format_cell_for_html(cell_value, col_idx)
                if isinstance(cell_value, (int, float)) and cell_value not in ("", None) and col_idx > 1:
                    cell_classes.append("currency")
                
                class_attr = " ".join(cell_classes) if cell_classes else ""
                attr_parts = []
                if class_attr:
                    attr_parts.append(f"class='{class_attr}'")
                attr_parts.append(f"data-cell=\"R{orig_row_idx}C{col_idx}\"")
                attr_parts.append(f"data-sheet=\"{actual_sheet_name}\"")

                is_cell_unlocked = False
                try:
                    cell_obj = sheet.cell(row=orig_row_idx, column=col_idx)
                    is_cell_unlocked = not bool(getattr(cell_obj.protection, 'locked', True))
                except Exception:
                    is_cell_unlocked = False

                if is_cell_unlocked and col_idx > 1:
                    attr_parts.append("data-editable=\"true\"")
                attr_str = " " + " ".join(attr_parts) if attr_parts else ""
                html_parts.append(f"    <td{attr_str}>{formatted_value}</td>")
            
            html_parts.append("  </tr>")
        
        html_parts.extend([
            "</tbody>",
            "</table>",
            "</div>",
            "</div>",
            "",
            "<!-- Footer Section -->",
            "<div class='report-footer'>",
            "<div class='footer-content'>",
            f"<h3 class='footer-title'><i class='fas fa-check-circle'></i> Report Generated Successfully</h3>",
            "</div>",
            "<div class='timestamp-badge'>",
            f"<i class='fas fa-clock'></i> Generated on {datetime.datetime.now().strftime('%B %d, %Y at %I:%M %p')}",
            "</div>",
            "</div>",
            "</div>",
            "",
            "</div>",
            "</div>",
            "",
            "<script>",
            "// Store JSON data for programmatic access",
            f"window.reportData = {json.dumps(json_data, ensure_ascii=False)};",
            "",
            "console.log('%c📊 Financial Report Data Loaded', 'color: #7c3aed; font-weight: bold; font-size: 16px; font-family: Inter, sans-serif;');",
            "console.log('%c━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━', 'color: #7c3aed;');",
            "console.log('%c📄 Sheet Name:', 'color: #6b7280; font-weight: 600;', window.reportData.sheetName);",
            "console.log('%c🔢 Total Cells:', 'color: #6b7280; font-weight: 600;', Object.keys(window.reportData.data).length);",
            "console.log('%c⏰ Timestamp:', 'color: #6b7280; font-weight: 600;', window.reportData.timestamp);",
            "console.log('%c━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━', 'color: #7c3aed;');",
            "console.log('%c💡 Access data: window.reportData.data[\"R1C1\"]', 'color: #10b981; font-style: italic;');",
            "",
            "// Download report as PDF (placeholder function)",
            "function downloadReport() {",
            "  alert('PDF download functionality will be implemented by the backend. (Ctrl+P to print)');",
            "  console.log('Download request initiated for:', window.reportData.sheetName);",
            "}",
            "",
            "// Add smooth scroll behavior",
            "document.querySelectorAll('a[href^=\"#\"]').forEach(anchor => {",
            "  anchor.addEventListener('click', function (e) {",
            "    e.preventDefault();",
            "    const target = document.querySelector(this.getAttribute('href'));",
            "    if (target) {",
            "      target.scrollIntoView({ behavior: 'smooth', block: 'start' });",
            "    }",
            "  });",
            "});",
            "",
            "// Add loading state handler",
            "window.addEventListener('load', () => {",
            "  document.querySelectorAll('.loading').forEach(el => {",
            "    el.classList.remove('loading');",
            "  });",
            "});",
            "",
            "// Add table row highlight on click",
            "document.querySelectorAll('tbody tr').forEach(row => {",
            "  row.addEventListener('click', function() {",
            "    document.querySelectorAll('tbody tr').forEach(r => {",
            "      r.style.outline = 'none';",
            "    });",
            "    this.style.outline = '2px solid var(--border-color)';",
            "    this.style.outlineOffset = '-2px';",
            "  });",
            "});",
            "",
            "// Add keyboard navigation",
            "document.addEventListener('keydown', (e) => {",
            "  if (e.ctrlKey && e.key === 'p') {",
            "    e.preventDefault();",
            "    window.print();",
            "  }",
            "});",
            "",
            "// Performance monitoring",
            "if (window.performance) {",
            "  const perfData = window.performance.timing;",
            "  const pageLoadTime = perfData.loadEventEnd - perfData.navigationStart;",
            "  console.log('%c⚡ Page Load Time:', 'color: #10b981; font-weight: 600;', pageLoadTime + 'ms');",
            "}",
            "",
            "// Add animation observer for elements",
            "const observerOptions = {",
            "  threshold: 0.1,",
            "  rootMargin: '0px 0px -50px 0px'",
            "};",
            "",
            "const observer = new IntersectionObserver((entries) => {",
            "  entries.forEach(entry => {",
            "    if (entry.isIntersecting) {",
            "      entry.target.style.opacity = '1';",
            "      entry.target.style.transform = 'translateY(0)';",
            "    }",
            "  });",
            "}, observerOptions);",
            "",
            "document.querySelectorAll('.table-wrapper').forEach(el => {",
            "  el.style.opacity = '0';",
            "  el.style.transform = 'translateY(20px)';",
            "  el.style.transition = 'opacity 0.6s ease, transform 0.6s ease';",
            "  observer.observe(el);",
            "});",
            "</script>",
            "</body>",
            "</html>"
        ])
        

        
        html_content = "\n".join(html_parts)
        print(f"[HTML Generator] SUCCESS: HTML generated using FALLBACK with professional styling ({len(html_content)} chars)", file=sys.stderr)
        print(f"[HTML Generator] SUCCESS: JSON data extracted ({len(json_data['data'])} cells)", file=sys.stderr)
        
        return html_content, json_data
        
    except Exception as e:
        print(f"[HTML Generator] ❌ Error generating HTML: {str(e)}", file=sys.stderr)
        import traceback
        traceback.print_exc(file=sys.stderr)
        return ""




def _abs_path(path: str) -> str:
    return os.path.abspath(path)


def _r1c1_to_a1(r1c1_ref: str) -> str:
    """
    Convert R1C1 notation (e.g., 'R3C2') to A1 notation (e.g., 'B3').
    
    Args:
        r1c1_ref: Cell reference in R1C1 format (e.g., 'R3C2')
        
    Returns:
        Cell reference in A1 format (e.g., 'B3')
    """
    import re
    match = re.match(r'R(\d+)C(\d+)', r1c1_ref, re.IGNORECASE)
    if not match:
        # If it's not R1C1 format, assume it's already A1 and return as-is
        return r1c1_ref
    
    row = int(match.group(1))
    col = int(match.group(2))
    
    # Convert column number to letter(s)
    col_letter = ''
    while col > 0:
        col -= 1
        col_letter = chr(65 + (col % 26)) + col_letter
        col //= 26
    
    return f'{col_letter}{row}'


def _collect_updates(workbook, updates: List[Dict[str, Any]]):
    from openpyxl.cell.cell import MergedCell
    
    print(f"[_collect_updates] Processing {len(updates)} updates", file=sys.stderr)
    
    applied = []
    for update in updates:
        sheet_name = update.get('sheet')
        cell_addr = update.get('cell')
        value = update.get('value')

        if not sheet_name or not cell_addr:
            continue

        if sheet_name not in workbook.sheetnames:
            raise ValueError(f'Sheet "{sheet_name}" not found in workbook')

        # Convert R1C1 notation to A1 if needed
        cell_addr_a1 = _r1c1_to_a1(cell_addr)

        sheet = workbook[sheet_name]
        cell = sheet[cell_addr_a1]
        
        # Debug logging for specific cells
        if cell_addr.lower() in ['i34', 'i35', 'h28', 'h30', 'h32', 'h33', 'h13', 'h14', 'h15']:
            print(f"[Update Debug] Cell {cell_addr} -> {cell_addr_a1} = {value}", file=sys.stderr)
        
        # Handle merged cells - write to the top-left cell of the merged range
        if isinstance(cell, MergedCell):
            # Find the merged range that contains this cell
            for merged_range in sheet.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    # Get the top-left cell of the merged range (this is the "master" cell)
                    top_left_cell = sheet.cell(merged_range.min_row, merged_range.min_col)
                    top_left_cell.value = value
                    print(f"[Update] Merged cell {cell_addr_a1} -> writing to master cell {top_left_cell.coordinate}", file=sys.stderr)
                    break
        else:
            # Normal cell - write directly
            cell.value = value

        applied.append({
            'sheet': sheet_name,
            'cell': cell_addr,  # Keep original format in response
            'value': value
        })

    print(f"[_collect_updates] Applied {len(applied)} updates successfully", file=sys.stderr)
    return applied


def _normalize_cell_reference(cell_addr: str) -> str:
    if not cell_addr:
        return ''
    return _r1c1_to_a1(cell_addr).upper()


def _apply_updates_via_com(excel_path: str, updates: List[Dict[str, Any]], save_as_path: str = None):
    if not COM_AVAILABLE:
        raise RuntimeError('COM automation is not available on this host')

    import win32com.client
    import pythoncom

    pythoncom.CoInitialize()
    excel_app = None
    wb_com = None
    applied_updates = []
    final_path = save_as_path or excel_path

    try:
        excel_app = win32com.client.DispatchEx("Excel.Application")
        try:
            excel_app.Visible = False
        except Exception:
            pass
        try:
            excel_app.DisplayAlerts = False
        except Exception:
            pass
        try:
            excel_app.AskToUpdateLinks = False
        except Exception:
            pass
        try:
            excel_app.ScreenUpdating = False
        except Exception:
            pass

        wb_com = excel_app.Workbooks.Open(excel_path, ReadOnly=True)
        
        # Get list of available sheet names for matching
        available_sheets = [wb_com.Worksheets(i).Name for i in range(1, wb_com.Worksheets.Count + 1)]
        print(f"[COM Update] Available sheets: {available_sheets}", file=sys.stderr)

        for update in updates or []:
            try:
                requested_sheet = update.get('sheet', wb_com.Worksheets(1).Name)
                cell_ref = _normalize_cell_reference(update.get('cell', ''))
                value = update.get('value')
                if not cell_ref:
                    continue
                
                # Find the matching sheet name (handles case and space differences)
                sheet_name = find_sheet_match(requested_sheet, available_sheets)
                if not sheet_name:
                    print(f"[COM Update] Sheet '{requested_sheet}' not found, trying direct access", file=sys.stderr)
                    sheet_name = requested_sheet
                
                ws = wb_com.Worksheets(sheet_name)
                ws.Range(cell_ref).Value = value
                applied_updates.append(update)
                print(f"[COM Update] {sheet_name}!{cell_ref} = {value}", file=sys.stderr)
            except Exception as update_err:
                print(f"[COM Update] Failed to update {update}: {update_err}", file=sys.stderr)

        try:
            wb_com.Application.Calculation = -4105  # xlCalculationAutomatic
        except Exception:
            pass
        wb_com.Application.CalculateFullRebuild()

        # Auto-fit columns on all sheets to prevent ######## display for numeric values (skip for index sheet)
        print(f"[COM Update] Auto-fitting columns on all sheets...", file=sys.stderr)
        for ws in wb_com.Worksheets:
            try:
                if normalize_sheet_name(ws.Name) != 'index':
                    ws.Columns.AutoFit()
            except Exception as autofit_err:
                print(f"[COM Update] Failed to AutoFit sheet '{ws.Name}': {autofit_err}", file=sys.stderr)

        if save_as_path:
            wb_com.SaveAs(save_as_path, FileFormat=51)
            final_path = save_as_path
        else:
            wb_com.Save()
            final_path = excel_path
    finally:
        if wb_com is not None:
            try:
                wb_com.Close(SaveChanges=False)
            except Exception:
                pass
        if excel_app is not None:
            try:
                excel_app.Quit()
            except Exception:
                pass
        pythoncom.CoUninitialize()

    return final_path, applied_updates


def calculate_excel(input_data: Dict[str, Any], excel_path: str) -> str:
    meta: Dict[str, Any] = {
        'templatePath': _abs_path(excel_path),
        'autoCalculation': 'enabled'  # Excel will auto-calculate formulas
    }

    try:
        # Check file extension
        file_ext = os.path.splitext(excel_path)[1].lower()
        
        if file_ext == '.xls':
            # For .xls files, prefer COM over pandas/openpyxl for better accuracy
            if COM_AVAILABLE and USE_COM_INTERFACE:
                print(f"[Excel Calculator] Loading .xls file with COM Excel automation: {excel_path}", file=sys.stderr)
                import tempfile

                output_dir = _get_safe_temp_dir()
                os.makedirs(output_dir, exist_ok=True)

                timestamp = datetime.datetime.now(datetime.UTC).strftime('%Y%m%dT%H%M%SZ')
                unique_id = str(uuid.uuid4())[:8]
                template_name = os.path.splitext(os.path.basename(excel_path))[0]
                output_path = _abs_path(
                    os.path.join(output_dir, f'{template_name}-updated-{timestamp}-{unique_id}.xlsx')
                )

                output_path, applied_updates = _apply_updates_via_com(
                    excel_path,
                    input_data.get('updates', []),
                    save_as_path=output_path
                )

                print(f"[Excel Calculator] ✓ .xls file processed with COM and saved as .xlsx", file=sys.stderr)

                # Load the saved .xlsx file for further processing
                workbook = openpyxl.load_workbook(output_path, data_only=True)
                
            else:
                # Fallback to pandas for .xls files when COM not available
                print(f"[Excel Calculator] COM not available, using pandas to convert .xls to .xlsx: {excel_path}", file=sys.stderr)
                import pandas as pd
                
                # Read all sheets from .xls file
                xls_data = pd.read_excel(excel_path, sheet_name=None, engine='xlrd')
                
                # Use TEMP_DIR environment variable, fallback to system temp directory
                import tempfile
                output_dir = _get_safe_temp_dir()
                os.makedirs(output_dir, exist_ok=True)
                
                timestamp = datetime.datetime.utcnow().strftime('%Y%m%dT%H%M%SZ')
                template_name = os.path.splitext(os.path.basename(excel_path))[0]
                temp_xlsx_path = _abs_path(
                    os.path.join(output_dir, f'{template_name}-converted-{timestamp}.xlsx')
                )
                
                # Write to .xlsx format using pandas
                with pd.ExcelWriter(temp_xlsx_path, engine='openpyxl') as writer:
                    for sheet_name, df in xls_data.items():
                        df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                print(f"[Excel Calculator] ✓ Converted .xls to .xlsx using pandas", file=sys.stderr)
                
                # Now load with openpyxl and continue normal processing
                workbook = openpyxl.load_workbook(temp_xlsx_path)
                applied_updates = _collect_updates(workbook, input_data.get('updates', []))
                
                output_path = _abs_path(
                    os.path.join(output_dir, f'{template_name}-updated-{timestamp}.xlsx')
                )
                
                workbook.save(output_path)
                
                # Try to recalculate formulas if COM becomes available
                print(f"[Excel Calculator] Attempting formula recalculation...", file=sys.stderr)
                try:
                    if COM_AVAILABLE and USE_COM_INTERFACE:
                        import win32com.client
                        excel_app = win32com.client.DispatchEx("Excel.Application")
                        try:
                            excel_app.Visible = False
                        except Exception:
                            pass
                        try:
                            excel_app.DisplayAlerts = False
                        except Exception:
                            pass
                        try:
                            excel_app.AskToUpdateLinks = False
                        except Exception:
                            pass
                        
                        wb_com = excel_app.Workbooks.Open(output_path)
                        try:
                            wb_com.Application.Calculation = -4105
                        except Exception:
                            pass
                        wb_com.Application.CalculateFullRebuild()
                        
                        # Auto-fit columns to prevent ######## display (skip for index sheet)
                        for ws in wb_com.Worksheets:
                            try:
                                if normalize_sheet_name(ws.Name) != 'index':
                                    ws.Columns.AutoFit()
                            except Exception:
                                pass
                                
                        wb_com.Save()
                        wb_com.Close(SaveChanges=True)
                        excel_app.Quit()
                        
                        print(f"[Excel Calculator] ✓ Formulas recalculated successfully", file=sys.stderr)
                except Exception as calc_error:
                    print(f"[Excel Calculator] ⚠ Formula recalculation failed: {calc_error}", file=sys.stderr)
                
                # Reload with data_only=True to strip formulas
                workbook.close()
                workbook = load_workbook(output_path, data_only=True)
                workbook.save(output_path)
        else:
            # Handle .xlsx files
            import tempfile
            output_dir = _get_safe_temp_dir()
            os.makedirs(output_dir, exist_ok=True)

            timestamp = datetime.datetime.utcnow().strftime('%Y%m%dT%H%M%SZ')
            unique_id = str(uuid.uuid4())[:8]
            template_name = os.path.splitext(os.path.basename(excel_path))[0]
            output_path = _abs_path(
                os.path.join(output_dir, f'{template_name}-updated-{timestamp}-{unique_id}.xlsx')
            )

            if COM_AVAILABLE and USE_COM_INTERFACE:
                print("[Excel Calculator] Applying .xlsx updates via COM automation", file=sys.stderr)
                output_path, applied_updates = _apply_updates_via_com(
                    excel_path,
                    input_data.get('updates', []),
                    save_as_path=output_path
                )
                workbook = openpyxl.load_workbook(output_path, data_only=True)
                print("[Excel Calculator] ✓ .xlsx file processed with COM", file=sys.stderr)
            else:
                print("[Excel Calculator] ⚠ COM unavailable, falling back to openpyxl", file=sys.stderr)
                workbook = openpyxl.load_workbook(excel_path)
                applied_updates = _collect_updates(workbook, input_data.get('updates', []))
                workbook.save(output_path)
                try:
                    workbook.close()
                except Exception:
                    pass
                
                # LIBREOFFICE FORMULA RECALCULATION (Linux Mode)
                if not USE_COM_INTERFACE:
                    try:
                        print("[Excel Calculator] triggering LibreOffice to recalculate formulas...", file=sys.stderr)
                        import subprocess
                        import shutil
                        
                        # Create a temp dir for the recalc output
                        recalc_dir = os.path.join(output_dir, f"recalc_{unique_id}")
                        os.makedirs(recalc_dir, exist_ok=True)
                        
                        # Run soffice to convert xlsx -> xlsx (forces recalc)
                        cmd = [
                            "soffice",
                            "--headless",
                            "--convert-to", "xlsx",
                            "--outdir", recalc_dir,
                            output_path
                        ]
                        
                        # Run with timeout
                        result = subprocess.run(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, timeout=60)
                        
                        if result.returncode == 0:
                            # Expected output filename
                            recalculated_filename = os.path.basename(output_path) # e.g. name.xlsx
                            recalculated_path = os.path.join(recalc_dir, recalculated_filename)
                            
                            if os.path.exists(recalculated_path):
                                # Replace the original output_path with the recalculated one
                                shutil.move(recalculated_path, output_path)
                                print("[Excel Calculator] ✓ LibreOffice formula recalculation successful", file=sys.stderr)
                            else:
                                print(f"[Excel Calculator] ⚠ Recalculated file not found at {recalculated_path}", file=sys.stderr)
                        else:
                            print(f"[Excel Calculator] ⚠ LibreOffice recalc failed: {result.stderr.decode()}", file=sys.stderr)
                            
                        # Cleanup
                        shutil.rmtree(recalc_dir, ignore_errors=True)
                        
                    except Exception as e:
                        print(f"[Excel Calculator] ⚠ Error during LibreOffice recalculation: {e}", file=sys.stderr)

                workbook = load_workbook(output_path, data_only=True)

        # Read the Excel file as bytes and encode to base64
        with open(output_path, 'rb') as f:
            excel_bytes = f.read()
        excel_base64 = base64.b64encode(excel_bytes).decode('utf-8')

        # Also extract JSON data for browser display in Luckysheet format (skip if requested)
        json_output = []
        if not input_data.get('skipJsonExtraction', False):
            # For Term Loan templates, use COM extraction for better accuracy
            template_upper = template_name.upper()
            use_com_extraction = 'TERM LOAN' in template_upper or 'TERM_LOAN' in template_upper
            
            if use_com_extraction and COM_AVAILABLE:
                print("[JSON Extraction] Using COM method for Term Loan template", file=sys.stderr)
                json_output = extract_sheet_data_with_com(output_path)
                if json_output is None:
                    print("[JSON Extraction] COM extraction failed, falling back to pandas", file=sys.stderr)
                    use_com_extraction = False
            
            if not use_com_extraction or json_output is None:
                # Fallback to pandas-based extraction
                print("[JSON Extraction] Using pandas method", file=sys.stderr)
                try:
                    import pandas as pd
                    all_sheets = pd.read_excel(output_path, sheet_name=None, engine='openpyxl')
                    json_output = []
                    for sheet_name, df in all_sheets.items():
                        try:
                            df_cleaned = df.replace([pd.NA, np.inf, -np.inf], None)
                            df_cleaned = df_cleaned.where(pd.notna(df_cleaned), None)

                            # Convert to Luckysheet format
                            sheet_data = []
                            max_rows = len(df_cleaned)
                            max_cols = len(df_cleaned.columns) if max_rows > 0 else 0
                            
                            for row_idx in range(max_rows):
                                row_data = []
                                for col_idx in range(max_cols):
                                    try:
                                        value = df_cleaned.iloc[row_idx, col_idx] if row_idx < len(df_cleaned) else None
                                        if value is not None and not pd.isna(value):
                                            cell_data = {
                                                'v': value,
                                                'm': str(value) if value is not None else ''
                                            }
                                        else:
                                            cell_data = None
                                        row_data.append(cell_data)
                                    except Exception as cell_error:
                                        print(f"Error processing cell {row_idx},{col_idx}: {cell_error}", file=sys.stderr)
                                        row_data.append(None)
                                sheet_data.append(row_data)
                            
                            sheet_obj = {
                                'name': sheet_name,
                                'data': sheet_data,
                                'config': {
                                    'merge': {},
                                    'borderInfo': [],
                                    'rowlen': {},
                                    'columnlen': {}
                                },
                                'index': len(json_output)  # sheet index
                            }
                            json_output.append(sheet_obj)
                        except Exception as sheet_error:
                            print(f"Error processing sheet {sheet_name}: {sheet_error}", file=sys.stderr)
                            # Skip this sheet
                            continue
                except Exception as json_error:
                    print(f"Error generating JSON data: {json_error}", file=sys.stderr)
                    json_output = []  # Fallback to empty array

        # Determine sheet name based on template format
        final_sheet_name = get_final_sheet_name(template_name)

        # ---------------------------------------------------------------------
        # Linux/LibreOffice Best Practice:
        # 1) (Optional) Recalculate formulas (handled earlier via COM or LibreOffice)
        # 2) Convert formulas -> cached values (flatten)
        # 3) Export PDFs from the flattened workbook
        # This prevents LibreOffice from re-evaluating formulas and producing #NAME?.
        # ---------------------------------------------------------------------
        pdf_source_path = output_path
        if (not USE_COM_INTERFACE) and FREEZE_FORMULAS_TO_VALUES and (
            (not input_data.get('skipPdf', False)) or input_data.get('generateFullReport', False)
        ):
            try:
                freeze_id = uuid.uuid4().hex[:8]
                flattened_pdf_source_path = _abs_path(
                    os.path.join(output_dir, f'{template_name}-flattened-{timestamp}-{freeze_id}.xlsx')
                )
                print(
                    f"[Excel Calculator] Flattening formulas to values for Linux PDF export: {flattened_pdf_source_path}",
                    file=sys.stderr,
                )
                stats = _freeze_workbook_formulas_to_values(output_path, flattened_pdf_source_path)
                print(
                    f"[Excel Calculator] ✓ Flattened workbook created (converted={stats.get('converted')}, missing_cached={stats.get('missing_cached')})",
                    file=sys.stderr,
                )
                pdf_source_path = flattened_pdf_source_path
                meta['flattenedPdfSource'] = pdf_source_path
                meta['flattenStats'] = stats
            except Exception as freeze_error:
                print(f"[Excel Calculator] ⚠ Could not flatten formulas for PDF export: {freeze_error}", file=sys.stderr)
                pdf_source_path = output_path

        # Generate PDF for Final workings sheet directly from Excel
        pdf_base64 = None
        pdf_file_name = None
        if not input_data.get('skipPdf', False):
            try:
                pdf_output_path = os.path.join(output_dir, f'{template_name}-{final_sheet_name}-{timestamp}.pdf')
                if generate_pdf_from_excel_sheet(pdf_source_path, final_sheet_name, pdf_output_path):
                    with open(pdf_output_path, 'rb') as f:
                        pdf_bytes = f.read()
                    pdf_base64 = base64.b64encode(pdf_bytes).decode('utf-8')
                    pdf_file_name = f'{template_name}-{final_sheet_name}-{timestamp}.pdf'
                    print(f"PDF generated successfully: {pdf_file_name}", file=sys.stderr)
                    # Clean up PDF file after encoding
                    os.unlink(pdf_output_path)
                else:
                    print("PDF generation failed", file=sys.stderr)
            except Exception as pdf_error:
                print(f"Error generating PDF: {pdf_error}", file=sys.stderr)
        else:
            print("PDF generation skipped as requested", file=sys.stderr)

        # Generate HTML for Final workings sheet with exact formatting (skip if requested)
        html_content = None
        html_json_data = {}
        if not input_data.get('skipHtmlGeneration', False):
            try:
                # Extract header data from input_data
                header_data = {
                    'proprietor': input_data.get('proprietor'),
                    'sector': input_data.get('sector'),
                    'natureOfBusiness': input_data.get('natureOfBusiness')
                }
                
                result_tuple = generate_html_from_excel_sheet(output_path, final_sheet_name, header_data=header_data)
                # Handle both old return (str) and new return (tuple) for backward compatibility
                if isinstance(result_tuple, tuple):
                    html_content, html_json_data = result_tuple
                else:
                    html_content = result_tuple
                    html_json_data = {}
                
                if html_content:
                    print(f"HTML generated successfully ({len(html_content)} chars)", file=sys.stderr)
                    if html_json_data:
                        print(f"HTML JSON data extracted ({len(html_json_data.get('data', {}))} cells)", file=sys.stderr)
                else:
                    print("HTML generation returned empty content", file=sys.stderr)
            except Exception as html_error:
                print(f"Error generating HTML: {html_error}", file=sys.stderr)
        else:
            print("HTML generation skipped as requested", file=sys.stderr)

        # Generate full AI-enhanced report if requested (Grok only)
        full_report_base64 = None
        full_report_filename = None
        if input_data.get('generateFullReport', False) and input_data.get('grokApiKey'):
            try:
                print(f"\n{'='*80}", file=sys.stderr)
                print(f"🚀 GENERATING FULL AI-ENHANCED REPORT", file=sys.stderr)
                print(f"{'='*80}\n", file=sys.stderr)
                
                # Import AI report generator
                from pdf_report_generator import AIReportGenerator
                
                # Create PDFs directory
                pdfs_dir = os.path.join(output_dir, f'pdfs_{timestamp}')
                os.makedirs(pdfs_dir, exist_ok=True)
                
                # Generate PDFs for all sheets
                print("[Full Report] Step 1: Generating PDFs for all Excel sheets...", file=sys.stderr)
                selected_sheets = input_data.get('selectedSheets')
                dynamic_excluded_sheets = input_data.get('excludedSheets')
                index_sheet_name = input_data.get('indexSheet')
                sheet_pdfs = generate_pdfs_for_all_sheets(
                    pdf_source_path,
                    pdfs_dir,
                    selected_sheets,
                    dynamic_excluded_sheets,
                    index_sheet_name
                )
                
                # Fallback: If no sheets were generated and selectedSheets was provided, try generating ALL sheets
                if sheet_pdfs.get('success_count', 0) == 0 and selected_sheets:
                    print("[Full Report] ⚠️  No sheets matched the selection. Falling back to generating ALL sheets.", file=sys.stderr)
                    sheet_pdfs = generate_pdfs_for_all_sheets(
                        pdf_source_path,
                        pdfs_dir,
                        None,  # No filter
                        dynamic_excluded_sheets,
                        index_sheet_name
                    )
                
                print(f"[Full Report] Generated {sheet_pdfs['success_count']} sheet PDFs", file=sys.stderr)

                requested_status_entries = sheet_pdfs.get('sheet_status') or []
                requested_total = len(requested_status_entries)
                if requested_total:
                    success_entries = [entry for entry in requested_status_entries if entry.get('status') == 'success']
                    failure_entries = [entry for entry in requested_status_entries if entry.get('status') != 'success']
                    print(
                        f"[Full Report] Sheet inclusion summary -> Requested: {requested_total}, "
                        f"Success: {len(success_entries)}, Failed: {len(failure_entries)}",
                        file=sys.stderr
                    )
                    for failure in failure_entries:
                        reason = failure.get('reason') or 'Unknown reason'
                        sheet_label = failure.get('sheet') or 'Unnamed sheet'
                        print(f"[Full Report]    ✗ {sheet_label}: {reason}", file=sys.stderr)
                else:
                    print(
                        "[Full Report] Sheet inclusion summary unavailable (no requested sheets specified)",
                        file=sys.stderr
                    )
                if sheet_pdfs.get('success_count', 0) == 0:
                    raise RuntimeError(
                        "Failed to generate any Excel sheet PDFs for the AI report. "
                        "Please review the LibreOffice/COM PDF generation logs for details."
                    )
                
                # Prepare Excel data for AI
                excel_data = {
                    'json_data': json_output,
                    'html_data': html_json_data,
                    'template_name': template_name,
                    'timestamp': timestamp
                }
                
                # Initialize AI generator - Grok only
                signature_path = input_data.get('signaturePath')
                ai_generator = AIReportGenerator(input_data['grokApiKey'], provider="grok", signature_path=signature_path)
                print("[Full Report] Using Grok AI for report generation", file=sys.stderr)                # Generate full report
                full_report_path = os.path.join(output_dir, f'{template_name}-full-report-{timestamp}.pdf')
                print("[Full Report] Step 2: Generating AI content and merging...", file=sys.stderr)
                
                report_result = ai_generator.generate_full_report(
                    excel_pdfs_dir=pdfs_dir,
                    excel_data=excel_data,
                    output_path=full_report_path,
                    template_name=template_name
                )
                
                if report_result['success']:
                    # Read and encode the full report
                    with open(full_report_path, 'rb') as f:
                        full_report_bytes = f.read()
                    full_report_base64 = base64.b64encode(full_report_bytes).decode('utf-8')
                    full_report_filename = os.path.basename(full_report_path)
                    
                    print(f"[Full Report] ✅ Full report generated: {full_report_filename}", file=sys.stderr)
                    print(f"[Full Report]    AI Sections: {len(report_result.get('ai_sections_generated', []))}", file=sys.stderr)
                    print(f"[Full Report]    Excel PDFs: {len(report_result.get('excel_pdfs_included', []))}", file=sys.stderr)
                    
                    # Clean up individual Excel sheet PDFs (only keep final report)
                    try:
                        import shutil
                        if os.path.exists(pdfs_dir):
                            shutil.rmtree(pdfs_dir)
                            print(f"[Full Report] 🗑️  Cleaned up individual sheet PDFs from {pdfs_dir}", file=sys.stderr)
                    except Exception as cleanup_error:
                        print(f"[Full Report] ⚠️  Could not clean up temp PDFs: {str(cleanup_error)}", file=sys.stderr)
                else:
                    print(f"[Full Report] ❌ Report generation failed", file=sys.stderr)
                    
            except Exception as full_report_error:
                print(f"[Full Report] Error generating full report: {str(full_report_error)}", file=sys.stderr)
                import traceback
                traceback.print_exc(file=sys.stderr)

        meta['verificationCopy'] = output_path

        result = {
            'success': True,
            'message': 'Workbook updated, encoded, PDF and HTML generated',
            '_appliedUpdates': applied_updates,
            '_meta': meta,
            'excelData': excel_base64,
            'jsonData': json_output,
            'pdfData': pdf_base64,
            'pdfFileName': pdf_file_name,
            'htmlContent': html_content,
            'htmlJsonData': html_json_data,  # Add extracted JSON data from HTML
            'fileName': f'{template_name}-updated-{timestamp}.xlsx',
            'fullReportData': full_report_base64,  # AI-enhanced full report
            'fullReportFileName': full_report_filename
        }

        return json.dumps(result, ensure_ascii=False, default=str)
    except Exception as exc:  # pragma: no cover - operational guard
        return json.dumps({'success': False, 'error': str(exc)})


if __name__ == '__main__':
    import sys
    import io
    
    # Force UTF-8 encoding for stdout to handle emojis and special characters
    if sys.stdout.encoding != 'utf-8':
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    
    args = sys.argv[1:]
    excel_file_path = args[0]
    json_input_arg = args[1]

    if json_input_arg.endswith('.json') and os.path.exists(json_input_arg):
        with open(json_input_arg, 'r', encoding='utf-8') as f:
            json_input_string = f.read()
    else:
         json_input_string = json_input_arg

    payload = json.loads(json_input_string)
    outcome = calculate_excel(payload, excel_file_path)
    print(outcome)
