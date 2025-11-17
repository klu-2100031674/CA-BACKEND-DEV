"""
Deep inspection of format CC2.xlsx to find all cell mapping issues
"""
import openpyxl
from pathlib import Path

# Load the workbook
wb_path = Path(__file__).parent.parent.parent / 'templates' / 'excel' / 'format CC2.xlsx'
wb = openpyxl.load_workbook(wb_path, data_only=False)

print("=" * 100)
print("DEEP INSPECTION OF FORMAT CC2.XLSX")
print("=" * 100)

# Check Assumptions.1 sheet
if 'Assumptions.1' in wb.sheetnames:
    ws = wb['Assumptions.1']
    
    print("\n" + "=" * 100)
    print("SECTION 1: GENERAL INFORMATION (Rows 3-8)")
    print("=" * 100)
    for row in range(3, 9):
        print(f"\nRow {row}:")
        for col in ['G', 'H', 'I']:
            cell = ws[f'{col}{row}']
            if cell.value:
                val = str(cell.value)
                if val.startswith('='):
                    print(f"  {col}{row}: FORMULA = {val[:80]}")
                else:
                    print(f"  {col}{row}: VALUE = {val}")
    
    print("\n" + "=" * 100)
    print("SECTION 2: MEANS OF FINANCE (Rows 10-14)")
    print("=" * 100)
    for row in range(10, 15):
        print(f"\nRow {row}:")
        for col in ['G', 'H', 'I']:
            cell = ws[f'{col}{row}']
            if cell.value:
                val = str(cell.value)
                if val.startswith('='):
                    print(f"  {col}{row}: FORMULA = {val[:80]}")
                else:
                    print(f"  {col}{row}: VALUE = {val}")
    
    print("\n" + "=" * 100)
    print("SECTION 3: FINANCIAL YEARS (Rows 16-20)")
    print("=" * 100)
    for row in range(16, 21):
        print(f"\nRow {row}:")
        for col in ['G', 'H', 'I']:
            cell = ws[f'{col}{row}']
            if cell.value:
                val = str(cell.value)
                if val.startswith('='):
                    print(f"  {col}{row}: FORMULA = {val[:80]}")
                else:
                    print(f"  {col}{row}: VALUE = {val}")
    
    print("\n" + "=" * 100)
    print("SECTION 4: AUDITED FINANCIAL STATEMENTS (Rows 22-43)")
    print("=" * 100)
    print("Checking for INPUT cells vs FORMULA cells...")
    input_cells = []
    formula_cells = []
    for row in range(22, 44):
        cell_i = ws[f'I{row}']
        if cell_i.value:
            val = str(cell_i.value)
            if val.startswith('='):
                formula_cells.append((row, val[:80]))
            else:
                input_cells.append((row, val))
    
    print(f"\nINPUT cells (should receive data from form):")
    for row, val in input_cells:
        print(f"  I{row}: {val}")
    
    print(f"\nFORMULA cells (auto-calculated, should NOT receive data):")
    for row, formula in formula_cells:
        print(f"  I{row}: {formula}")
    
    print("\n" + "=" * 100)
    print("SECTION 5: PROVISIONAL STATEMENTS (Rows 44-66)")
    print("=" * 100)
    input_cells_prov = []
    formula_cells_prov = []
    for row in range(44, 67):
        cell_i = ws[f'I{row}']
        if cell_i.value:
            val = str(cell_i.value)
            if val.startswith('='):
                formula_cells_prov.append((row, val[:80]))
            else:
                input_cells_prov.append((row, val))
    
    print(f"\nINPUT cells (should receive data from form):")
    for row, val in input_cells_prov:
        print(f"  I{row}: {val}")
    
    print(f"\nFORMULA cells (auto-calculated, should NOT receive data):")
    for row, formula in formula_cells_prov:
        print(f"  I{row}: {formula}")
    
    print("\n" + "=" * 100)
    print("SECTION 6: ASSUMPTIONS (Rows 67-99)")
    print("=" * 100)
    
    # Check specific assumption rows
    critical_rows = [67, 68, 69, 70, 71, 72, 73, 74, 75, 76, 77, 78, 79, 80, 81, 82, 83, 84, 85, 86, 87, 88, 89, 90, 91, 92, 93, 94, 95, 96, 97, 98, 99]
    
    for row in critical_rows:
        print(f"\nRow {row}:")
        for col in ['G', 'H', 'I']:
            cell = ws[f'{col}{row}']
            if cell.value:
                val = str(cell.value)
                if val.startswith('='):
                    print(f"  {col}{row}: FORMULA = {val[:100]}")
                else:
                    print(f"  {col}{row}: VALUE = {val}")
    
    print("\n" + "=" * 100)
    print("SECTION 7: FIXED ASSETS - Checking where to write D/E columns")
    print("=" * 100)
    
    # Check Fixed Assets rows (100, 110, 120, etc.)
    fixed_asset_rows = [100, 110, 120, 130, 133, 142, 152, 162, 172, 181, 191]
    
    for row in fixed_asset_rows:
        print(f"\nRow {row} (Fixed Asset Category):")
        for col in ['D', 'E']:
            cell = ws[f'{col}{row}']
            if cell.value:
                val = str(cell.value)
                if val.startswith('='):
                    print(f"  {col}{row}: FORMULA = {val[:80]}")
                else:
                    print(f"  {col}{row}: VALUE = {val}")
            else:
                print(f"  {col}{row}: EMPTY (ready for input)")
    
    print("\n" + "=" * 100)
    print("CRITICAL ISSUES SUMMARY")
    print("=" * 100)
    
    # List all cells with formulas that we might be overwriting
    print("\n⚠️  CELLS WITH FORMULAS IN COLUMN I (rows 1-99):")
    for row in range(1, 100):
        cell = ws[f'I{row}']
        if cell.value and str(cell.value).startswith('='):
            print(f"  I{row}: {str(cell.value)[:100]}")
    
    print("\n⚠️  CELLS WITH FORMULAS IN COLUMN H (rows 1-99):")
    for row in range(1, 100):
        cell = ws[f'H{row}']
        if cell.value and str(cell.value).startswith('='):
            print(f"  H{row}: {str(cell.value)[:100]}")

else:
    print("❌ Assumptions.1 sheet not found!")
    print("Available sheets:", wb.sheetnames)

# Check Final workings sheet to see what it references
if 'Final workings' in wb.sheetnames or 'Finalworkings' in wb.sheetnames:
    print("\n" + "=" * 100)
    print("FINAL WORKINGS SHEET - Sample cells to understand structure")
    print("=" * 100)
    
    final_sheet_name = 'Final workings' if 'Final workings' in wb.sheetnames else 'Finalworkings'
    ws_final = wb[final_sheet_name]
    
    # Check where Sector and Nature of Business are coming from
    print("\nChecking cells that display in the output report:")
    
    # Try to find Sector field
    for row in range(1, 30):
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            cell = ws_final[f'{col}{row}']
            if cell.value and 'Sector' in str(cell.value):
                print(f"\n'{cell.value}' found at {col}{row}")
                # Check adjacent cells
                for adj_col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
                    adj_cell = ws_final[f'{adj_col}{row}']
                    if adj_cell.value:
                        val = str(adj_cell.value)
                        print(f"  {adj_col}{row}: {val[:100]}")

wb.close()
